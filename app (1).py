# -*- coding: utf-8 -*-
"""
OFX Bridge — Interface Streamlit
Convertisseur de relevés bancaires PDF vers OFX
Supporte : Qonto, LCL, CA, CE, BP, CIC, CGD, LBP, SG, BNP, myPOS, Shine,
           CBAO, Ecobank, BCI, Coris, UBA, Orabank, BOA, ATB, BSIC, BIS, BNDE
"""

import io
import re
import hashlib
import logging
import tempfile
import os
from datetime import datetime
from pathlib import Path

import streamlit as st

# ── Logging minimal (Streamlit gère son propre stdout) ───────────────────────
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(funcName)s — %(message)s")
logger = logging.getLogger("ofxbridge")

# ── Imports avec gestion d'erreur propre ─────────────────────────────────────
try:
    import pdfplumber
    _PDFPLUMBER_OK = True
except ImportError:
    _PDFPLUMBER_OK = False

_OCR_AVAILABLE = False
_EASYOCR_OK = False
_TESSERACT_OK = False

try:
    import easyocr
    _EASYOCR_OK = True
    _OCR_AVAILABLE = True
except ImportError:
    pass

if not _EASYOCR_OK:
    try:
        import pytesseract
        _TESSERACT_OK = True
        _OCR_AVAILABLE = True
    except ImportError:
        pass

try:
    from pdf2image import convert_from_path
    _PDF2IMAGE_OK = True
except ImportError:
    _PDF2IMAGE_OK = False

try:
    from pydantic import BaseModel, field_validator
    _PYDANTIC_OK = True
except ImportError:
    _PYDANTIC_OK = False

# ── OCR via PyMuPDF (fitz) — fallback sans Tesseract ─────────────────────────
_FITZ_OK = False
try:
    import fitz  # PyMuPDF
    _FITZ_OK = True
except ImportError:
    pass

import base64
import json
import urllib.request

# ════════════════════════════════════════════════════════════════════════════
# MODÈLE PYDANTIC
# ════════════════════════════════════════════════════════════════════════════
if _PYDANTIC_OK:
    class Transaction(BaseModel):
        date:   str
        type:   str
        amount: float
        name:   str
        memo:   str = ""
        fitid:  str

        @field_validator('date')
        @classmethod
        def date_must_be_8_digits(cls, v):
            if not re.match(r'^\d{8}$', v):
                raise ValueError(f"Date OFX invalide: '{v}'")
            return v

        @field_validator('type')
        @classmethod
        def type_must_be_valid(cls, v):
            if v not in ('CREDIT', 'DEBIT'):
                raise ValueError(f"Type invalide: '{v}'")
            return v

        @field_validator('amount')
        @classmethod
        def amount_must_be_nonzero(cls, v):
            if v == 0.0:
                raise ValueError("Montant nul détecté")
            return v


# ════════════════════════════════════════════════════════════════════════════
# UTILITAIRES COMMUNS (identiques à la version Tkinter)
# ════════════════════════════════════════════════════════════════════════════

def extract_words_by_page(pdf_path):
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_words(keep_blank_chars=False))
    return pages

def extract_text_by_page(pdf_path):
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    return pages

def parse_amount(s):
    """
    Parse un montant depuis une chaîne.
    Supporte : EUR (1 234,56 / 1.234,56 / 1,234.56), XOF entiers (1 234 567),
    formats mixtes OCR, espaces insécables.
    """
    s = re.sub(r'\s+', ' ', str(s))  # normalise \n \t \r en espace
    s = s.replace('\xa0','').replace('\u202f','').replace(' ','').replace('*','').strip()
    # Format européen : 1.234,56
    if re.match(r'^\d{1,3}(\.\d{3})+,\d{2}$', s):
        return float(s.replace('.','').replace(',','.'))
    # Format simple virgule : 1234,56
    if re.match(r'^\d+,\d{2}$', s):
        return float(s.replace(',','.'))
    # Format point décimal anglais : 1234.56
    if re.match(r'^\d+\.\d{2}$', s):
        return float(s)
    # XOF entier pur (sans décimale) : ex "2100000" issu d'un token sans espace
    if re.match(r'^\d{4,}$', s):
        return float(s)
    # Nettoyage générique
    cleaned = re.sub(r'[^\d,.]', '', s)
    # Si plusieurs virgules ou points → essayer de déduire le séparateur
    dots = cleaned.count('.')
    commas = cleaned.count(',')
    if dots == 1 and commas == 0:
        try: return float(cleaned)
        except ValueError: pass
    if commas == 1 and dots == 0:
        try: return float(cleaned.replace(',', '.'))
        except ValueError: pass
    # Format ambigu : garder la dernière virgule/point comme décimale
    cleaned = cleaned.replace(',', '.')
    parts = cleaned.split('.')
    if len(parts) >= 2:
        integer_part = ''.join(parts[:-1])
        decimal_part = parts[-1]
        try: return float(f"{integer_part}.{decimal_part}")
        except ValueError: pass
    try:
        return float(cleaned)
    except ValueError:
        return None

def group_words_by_row(words, tol=3.0):
    if not words:
        return []
    rows, cur, top = [], [words[0]], words[0]['top']
    for w in words[1:]:
        if abs(w['top'] - top) <= tol:
            cur.append(w)
        else:
            rows.append(sorted(cur, key=lambda x: x['x0']))
            cur, top = [w], w['top']
    if cur:
        rows.append(sorted(cur, key=lambda x: x['x0']))
    return sorted(rows, key=lambda r: r[0]['top'])

def clean_label(s):
    return re.sub(r'\s+', ' ', s).strip()

def _is_technical_label(label):
    if not label:
        return True
    if re.match(r'^\d{6}\s+CB\*+\d+\s+\w+\s*$', label):
        return True
    if not re.search(r'[A-Za-zÀ-ÿ]{3,}', label):
        return True
    return False

def _is_human_readable(label):
    if not label:
        return False
    if re.search(r'[A-Z0-9]{15,}', label):
        return False
    if re.match(r'^[\d\s\-\/.,]+$', label):
        return False
    readable_words = [w for w in label.split() if re.search(r'[A-Za-zÀ-ÿ]{2,}', w)
                      and not re.match(r'^\d', w)]
    return len(readable_words) >= 2

def smart_label(main_label, memo_lines):
    label = clean_label(main_label)
    memos = [clean_label(m) for m in memo_lines if clean_label(m)]
    if _is_technical_label(label) and memos:
        for candidate in memos:
            if _is_human_readable(candidate):
                remaining = ' | '.join(m for m in memos if m != candidate and m)
                return candidate, (label + (' | ' + remaining if remaining else ''))
        return label, ' | '.join(memos)
    return label, ' | '.join(memos)

def make_fitid(date, label, amount):
    return hashlib.md5(f"{date}{label}{amount:.2f}".encode()).hexdigest()

def date_jjmm_to_ofx(jjmm, year):
    p = jjmm.replace('.', '/').split('/')
    if len(p) == 2:
        return f"{year}{p[1].zfill(2)}{p[0].zfill(2)}"
    return f"{year}0101"

def date_full_to_ofx(date_str):
    date_str = date_str.replace('.', '/')
    p = date_str.split('/')
    if len(p) == 3:
        return f"{p[2]}{p[1].zfill(2)}{p[0].zfill(2)}"
    return datetime.now().strftime('%Y%m%d')

def extract_iban(text):
    """
    Extrait le premier IBAN valide trouvé dans le texte.
    Stratégies (ordre de priorité) :
      1. Mot-clé IBAN + lecture ligne par ligne (évite la fusion avec le BIC)
      2. IBAN nu — format groupé par 4 sans mot-clé
      3. Fallback UEMOA/BCEAO sans mot-clé
    Tronque à la longueur IBAN réglementaire du pays (ex: FR=27, SN=28…).
    """
    # Longueurs IBAN officielles par code pays
    _IBAN_LEN = {
        'FR':27,'BE':16,'DE':22,'ES':24,'IT':27,'NL':18,'PT':25,'GB':22,
        'CH':21,'IE':22,'LU':20,'AT':20,'DK':18,'FI':18,'NO':15,'SE':24,
        'SN':28,'CI':28,'BJ':28,'TG':28,'ML':28,'BF':28,'NE':28,
        'GW':25,'GN':26,'MR':27,'CM':27,'MA':28,'TN':24,'DZ':24,
    }

    def _clean_and_truncate(raw):
        """Nettoie et tronque à la bonne longueur selon le code pays."""
        raw = re.sub(r'\s+', '', raw).upper()
        raw = re.sub(r'[^A-Z0-9]', '', raw)
        if len(raw) < 14 or not re.match(r'^[A-Z]{2}\d{2}', raw):
            return ''
        max_len = _IBAN_LEN.get(raw[:2], 34)
        return raw[:max_len]

    # Normaliser les espaces insécables
    text_clean = text.replace('\xa0', ' ').replace('\u202f', ' ')

    # ── 1. Avec mot-clé : traiter ligne par ligne pour éviter la fusion avec BIC ─
    # Le BIC est souvent sur la ligne suivante : "IBAN: FR76 ...\nBIC: QNTO..."
    for line in text_clean.split('\n'):
        line = line.strip()
        m = re.search(
            r'(?:IBAN|I\.?B\.?A\.?N\.?)\s*[:\-]?\s*'
            r'([A-Z]{2}[\s]?\d{2}[\s\dA-Z]{10,38})',
            line, re.IGNORECASE
        )
        if m:
            result = _clean_and_truncate(m.group(1))
            if result:
                return result

    # ── 2. IBAN nu sans mot-clé — format groupé par 4 (ex: FR76 3000 4028…) ───
    # On cherche ligne par ligne pour ne pas déborder sur la ligne suivante
    for line in text_clean.split('\n'):
        m2 = re.search(
            r'\b([A-Z]{2}\d{2}(?:\s?[A-Z0-9]{4}){3,7}(?:\s?[A-Z0-9]{1,4})?)\b',
            line.upper()
        )
        if m2:
            result = _clean_and_truncate(m2.group(1))
            if result:
                return result

    # ── 3. IBAN UEMOA/BCEAO avec mot-clé (peut contenir lettres dans le BBAN) ──
    # Ex : "IBAN : SN08 SN035 01010 0022015022 03" ou compact
    uemoa_cc = r'(?:SN|CI|BJ|TG|ML|BF|NE|GW|GN|MR|CM|MA|TN|DZ)'
    for line in text_clean.split('\n'):
        m3a = re.search(
            r'(?:IBAN|I\.?B\.?A\.?N\.?)\s*[:\-]?\s*'
            r'(' + uemoa_cc + r'[\s\dA-Z]{14,36})',
            line, re.IGNORECASE
        )
        if m3a:
            result = _clean_and_truncate(m3a.group(1))
            if result:
                return result

    # ── 4. Fallback UEMOA/BCEAO sans mot-clé (scan global) ───────────────────
    m4 = re.search(
        r'\b(' + uemoa_cc + r'\d{2}[A-Z0-9\s]{15,35})\b',
        text_clean.upper()
    )
    if m4:
        result = _clean_and_truncate(m4.group(1))
        if result:
            return result

    # ── 5. Numéro de compte RIB brut BCEAO (si aucun IBAN trouvé) ────────────
    # Format BCEAO numérique pur : ex "SN011 01005 005000458982 90"
    # Certains PDF africains écrivent le RIB sans mentionner "IBAN"
    m5 = re.search(
        r'\b([A-Z]{2}\d{3,5})\s+(\d{4,6})\s+(\d{8,14})\s+(\d{2})\b',
        text_clean.upper()
    )
    if m5:
        raw = ''.join(m5.groups())
        if len(raw) >= 14 and re.match(r'^[A-Z]{2}', raw):
            result = _clean_and_truncate(raw)
            if result:
                return result

    return ''

# Codes pays UEMOA → (longueur IBAN, longueur compte)
_UEMOA_IBAN = {
    # Format BCEAO : CC KK BBB AAAAA NNNNNNNNNNNN CC
    # CC=pays(2) KK=clé(2) BBB=banque(3-5) AAAAA=agence(4-5) N=compte(11-12) CC=clé RIB(2)
    'SN': 28,  # Sénégal
    'CI': 28,  # Côte d'Ivoire
    'BJ': 28,  # Bénin
    'TG': 28,  # Togo
    'ML': 28,  # Mali
    'BF': 28,  # Burkina Faso
    'NE': 28,  # Niger
    'GW': 25,  # Guinée-Bissau
    'GN': 26,  # Guinée
    'MR': 27,  # Mauritanie
    'CM': 27,  # Cameroun
    'MA': 28,  # Maroc
    'TN': 24,  # Tunisie
    'DZ': 24,  # Algérie
}

def iban_to_rib(iban, info=None):
    """
    Décompose un IBAN en (banque, agence, compte) pour l'OFX.
    Priorité : champs _rib_* extraits directement du PDF (via _afr_header).
    Sinon : France (FR27), UEMOA/BCEAO (SN28…), fallback numérique.
    """
    # ── Priorité : RIB directement extrait du PDF ────────────────────────────
    if info and info.get('_rib_bank') and info.get('_rib_account'):
        return (info['_rib_bank'],
                info.get('_rib_agency', '00000'),
                info['_rib_account'])

    c = iban.replace(' ', '').upper()
    c = re.sub(r'[^A-Z0-9]', '', c)   # purge tout caractère invalide

    # ── France ──────────────────────────────────────────────────────────────
    if c.startswith('FR') and len(c) == 27:
        r = c[4:]
        return r[0:5], r[5:10], r[10:21]

    # ── UEMOA/BCEAO ─────────────────────────────────────────────────────────
    country = c[:2]
    if country in _UEMOA_IBAN and len(c) >= 20:
        bban = c[4:]   # tout après CC+KK
        # Format BCEAO : BB(5 alphan.) + Agence(5 num.) + Compte(11-12 num.) + [Cle(2)]
        # Certains IBAN BSIC font 26 chars sans cle RIB -> ne pas tronquer avec [-2]
        code_banque = bban[0:5]
        agence      = bban[5:10]
        expected_len = _UEMOA_IBAN.get(country, 28)
        iban_has_key = len(c) >= expected_len
        compte_raw = bban[10:-2] if iban_has_key else bban[10:]
        compte     = re.sub(r'[A-Z]', '', compte_raw)
        return code_banque, agence, compte

    # ── Fallback numérique ───────────────────────────────────────────────────
    # Si on a un code pays UEMOA mais que l'IBAN était trop court pour le bloc
    # ci-dessus, on retente avec le BBAN brut (lettres conservées pour code banque)
    if country in _UEMOA_IBAN and len(c) >= 12:
        bban = c[4:]
        code_banque = bban[0:5]
        agence      = re.sub(r'[^0-9]', '', bban[5:10]) or '00000'
        expected_len = _UEMOA_IBAN.get(country, 28)
        iban_has_key = len(c) >= expected_len
        compte_raw  = bban[10:-2] if iban_has_key else bban[10:]
        compte      = re.sub(r'[A-Z]', '', compte_raw)
        if code_banque:
            return code_banque, agence, compte
    digits = re.sub(r'[^0-9]', '', c)
    if len(digits) >= 15:
        return digits[0:5], digits[5:10], digits[10:22]
    if len(digits) >= 5:
        return digits[0:5], '00000', digits[5:]
    return '00000', '00000', c[:20] if c else '00000'

def _year_from_text(text):
    m = re.search(r'\b(20\d{2})\b', text)
    return int(m.group(1)) if m else datetime.now().year

def _parse_col_amount(words):
    if not words:
        return None
    full = ' '.join(w['text'] for w in words).replace('\xa0', ' ').strip()
    if full in ('.', ',', ''):
        return None
    m = re.search(r'(\d{1,3}(?:[.\s]\d{3})+,\d{2})', full)
    if m:
        val = parse_amount(m.group(1).replace(' ', '.'))
        if val is not None and val > 0:
            return val
    m2 = re.search(r'(\d+,\d{2})', full)
    if m2:
        val = parse_amount(m2.group(1))
        if val is not None and val > 0:
            return val
    return None

def _parse_signed_amount(words):
    if not words:
        return None
    full = ' '.join(w['text'] for w in words).replace('\xa0', ' ').strip()
    m = re.search(r'([+\-])\s*([\d\s]+[,.][\d]{2})', full)
    if m:
        sign = 1.0 if m.group(1) == '+' else -1.0
        val = parse_amount(m.group(2))
        if val is not None:
            return sign * val
    m2 = re.search(r'([\d\s]+[,.][\d]{2})', full)
    if m2:
        val = parse_amount(m2.group(1))
        if val is not None:
            return val
    return None

def _make_txn(date_ofx, amount, label, memo=''):
    txn_dict = {
        'date':   date_ofx,
        'type':   'CREDIT' if amount >= 0 else 'DEBIT',
        'amount': amount,
        'name':   clean_label(label)[:64],
        'memo':   clean_label(memo)[:128],
        'fitid':  make_fitid(date_ofx, label, amount)
    }
    if _PYDANTIC_OK:
        try:
            Transaction(**txn_dict)
        except Exception as exc:
            logger.warning("Transaction ignorée [%s | %s | %.2f] : %s",
                           date_ofx, label[:40], amount, exc)
            return None
    return txn_dict


def audit_transactions(txns):
    """
    Audit qualité des transactions extraites.
    Retourne une liste d'anomalies (label court, montant nul/suspect, date invalide).
    """
    issues = []
    for t in txns:
        if abs(t.get('amount', 0)) < 0.01:
            issues.append(f"Montant nul/suspect : {t.get('date','')} | {t.get('name','')[:40]}")
        if len(t.get('name', '')) < 3:
            issues.append(f"Libellé trop court : {t.get('date','')} | '{t.get('name','')}'")
        if not re.match(r'^\d{8}$', str(t.get('date', ''))):
            issues.append(f"Date invalide : '{t.get('date','')}' | {t.get('name','')[:30]}")
    return issues


def _parse_amount_xof(text):
    """
    Parse spécifique XOF : gère les entiers groupés par espaces ou virgules
    (ex: '1 234 567' ou '1,234,567') en plus des formats EUR standards.
    """
    if not text:
        return None
    text = str(text).replace('\xa0', ' ').replace('\u202f', ' ').strip()
    # Entier groupé par espaces : "2 100 000"
    if re.match(r'^\d{1,3}( \d{3})+$', text):
        return float(text.replace(' ', ''))
    # Entier groupé par virgules sans décimale : "2,100,000"
    if re.match(r'^\d{1,3}(,\d{3})+$', text):
        return float(text.replace(',', ''))
    # Fallback parse_amount standard
    return parse_amount(text)

def _detect_pdf_type(pages_text):
    """
    Détection robuste scan vs PDF texte.
    Retourne 'SCAN' ou 'TEXT'.
    Critères : quantité de texte, ratio non-ASCII, densité alphanumérique.
    """
    total_chars = sum(len(p.strip()) for p in pages_text)

    # Trop peu de texte → clairement un scan
    if total_chars < 30:
        return "SCAN"

    # Ratio de caractères non-ASCII élevé → artefacts de scan
    non_ascii = sum(len(re.findall(r'[^\x00-\x7F]', p)) for p in pages_text)
    weird_ratio = non_ascii / max(total_chars, 1)
    if weird_ratio > 0.3:
        return "SCAN"

    # Pas assez de texte substantiel (min 300 chars avec densité alpha OK)
    all_text = ' '.join(pages_text)
    alpha_count = sum(1 for c in all_text if c.isalpha())
    if total_chars < 300 or (len(all_text) > 0 and alpha_count / len(all_text) < 0.15):
        return "SCAN"

    return "TEXT"


def _pdf_has_text(pages_text, min_chars=300):
    """Wrapper de compatibilité → utilise _detect_pdf_type."""
    return _detect_pdf_type(pages_text) == "TEXT"


def _preprocess_image_for_ocr(img):
    """
    Prétraitement optimisé pour relevés bancaires.
    Pipeline : niveaux de gris → MedianFilter (débruitage) → contraste → netteté → binarisation.
    """
    try:
        from PIL import ImageFilter, ImageEnhance
        import numpy as np

        # Niveaux de gris
        gray = img.convert('L')

        # Débruitage (MedianFilter recommandé pour documents texte)
        gray = gray.filter(ImageFilter.MedianFilter(size=3))

        # Contraste fort (2.5 optimal pour chiffres bancaires)
        gray = ImageEnhance.Contrast(gray).enhance(2.5)

        # Netteté
        gray = ImageEnhance.Sharpness(gray).enhance(2.0)

        # Binarisation adaptative
        arr = np.array(gray)
        threshold = arr.mean() * 0.85
        binary = (arr > threshold).astype(np.uint8) * 255
        from PIL import Image
        return Image.fromarray(binary)
    except Exception:
        return img


@st.cache_resource(show_spinner=False)
def _get_easyocr_reader():
    """Cache du reader EasyOCR (chargement ~10s la première fois)."""
    return easyocr.Reader(['fr', 'en'], gpu=False)


def _fix_ocr_common_errors(text):
    """
    Corrige les erreurs OCR classiques sur les relevés bancaires.
    | → 1 (séparateur confondu avec chiffre), O majuscule → 0 dans les montants.
    """
    # '|' confondu avec '1' dans les colonnes de montants
    text = text.replace('|', '1')
    # 'O' majuscule → '0' uniquement dans des séquences numériques
    text = re.sub(r'(?<=\d)O(?=\d)', '0', text)
    text = re.sub(r'\bO(\d)', r'0\1', text)
    text = re.sub(r'(\d)O\b', r'\g<1>0', text)
    return text


def _ocr_pdf(pdf_path):
    """
    OCR robuste : EasyOCR (priorité) avec preprocessing + corrections automatiques,
    fallback Tesseract si EasyOCR absent.
    """
    if not _PDF2IMAGE_OK:
        raise RuntimeError("pdf2image non installé — impossible de convertir le PDF en images.")
    if not _OCR_AVAILABLE:
        raise RuntimeError("Ce PDF semble scanné. Installez easyocr ou pytesseract+Tesseract.")

    images = convert_from_path(pdf_path, dpi=300)
    pages_text = []

    for img in images:
        processed = _preprocess_image_for_ocr(img)

        if _EASYOCR_OK:
            try:
                reader = _get_easyocr_reader()
                import numpy as np
                arr = np.array(processed)
                results = reader.readtext(arr, detail=1, paragraph=False,
                                          width_ths=0.7, add_margin=0.05)
                # Trier par Y (ligne) puis X (colonne) pour préserver la structure
                results_sorted = sorted(results, key=lambda r: (r[0][0][1], r[0][0][0]))
                lines, prev_y = [], -999
                for bbox, text, conf in results_sorted:
                    y = bbox[0][1]
                    # Corrections OCR communes
                    text = _fix_ocr_common_errors(text)
                    if y - prev_y > 15:
                        lines.append(text)
                    else:
                        lines[-1] = lines[-1] + ' ' + text
                    prev_y = y
                pages_text.append('\n'.join(lines))
            except Exception as exc:
                logger.warning("EasyOCR échoué : %s — fallback Tesseract", exc)
                if _TESSERACT_OK:
                    raw = pytesseract.image_to_string(processed, lang='fra+eng')
                    pages_text.append(_fix_ocr_common_errors(raw))
                else:
                    pages_text.append('')
        elif _TESSERACT_OK:
            raw = pytesseract.image_to_string(processed, lang='fra+eng')
            pages_text.append(_fix_ocr_common_errors(raw))
        else:
            pages_text.append('')

    return pages_text


# ════════════════════════════════════════════════════════════════════════════
# DÉTECTION DE LA BANQUE
# ════════════════════════════════════════════════════════════════════════════

def detect_bank(pages_text):
    text = pages_text[0][:3000].upper()
    if 'QONTO' in text or 'QNTOFRP' in text:
        return 'QONTO'
    if 'CREDIT LYONNAIS' in text or ('LCL' in text and 'RELEVE DE COMPTE COURANT' in text):
        return 'LCL'
    text_nospace = text.replace(' ', '')
    if ('SOCIETE GENERALE' in text or 'SOCIÉTÉ GÉNÉRALE' in text
            or 'SOCIETEGENERALE' in text_nospace) and (
            'SENEGAL' in text or 'SÉNÉGAL' in text or 'COTE D' in text
            or "CÔTE D'" in text or 'CAMEROUN' in text or 'DAKAR' in text
            or 'ABIDJAN' in text or 'DOUALA' in text or 'LOME' in text
            or 'BAMAKO' in text):
        return 'SG_AFRIQUE'
    if ('SOCIETE GENERALE' in text or 'SOCIÉTÉ GÉNÉRALE' in text
            or '552 120 222' in text or 'SOCIETEGENERALE' in text_nospace
            or 'SG.FR' in text or 'PROFESSIONNELS.SG.FR' in text):
        return 'SG'
    if 'CREDIT AGRICOLE' in text or 'AGRIFRPP' in text:
        return 'CA'
    if 'CAIXA GERAL' in text or 'CGDIFRPP' in text or 'CGD' in text[:500]:
        return 'CGD'
    if "CAISSE D'EPARGNE" in text or "CAISSE D.EPARGNE" in text or 'CEPAFRPP' in text:
        return 'CE'
    if 'BANQUE POPULAIRE' in text or 'CCBPFRPP' in text:
        return 'BP'
    if 'BANQUE POSTALE' in text or 'PSSTFRPP' in text or 'LABANQUEPOSTALE' in text:
        return 'LBP'
    if 'CREDIT INDUSTRIEL' in text or 'CMCIFRPP' in text or ('CIC' in text and 'RELEVE' in text):
        return 'CIC'
    if ('BNP PARIBAS' in text or 'BNPAFRPP' in text or 'BNP' in text[:500]
            or 'BANQUE NATIONALE DE PARIS' in text):
        return 'BNP'
    if 'MYPOS' in text or 'MYPOS LTD' in text or 'MY POS' in text:
        return 'MYPOS'
    if ('SNNNFR22XXX' in text or 'SHINE.FR' in text or 'SHINE SAS' in text
            or ('SHINE' in text and ('RELEVE' in text or 'SNNN' in text or '1741' in text))):
        return 'SHINE'
    if 'CBAO' in text or 'COMPAGNIE BANCAIRE DE L' in text:
        return 'CBAO'
    if 'ECOBANK' in text or 'ECOBANK SENEGAL' in text:
        return 'ECOBANK'
    if 'BANQUE POUR LE COMMERCE' in text and 'INDUSTRIE' in text:
        return 'BCI'
    if 'CORIS BANK' in text or 'CORISBANK' in text_nospace:
        return 'CORIS'
    if 'UNITED BANK FOR AFRICA' in text or 'UNAFSNDA' in text or ('UBA' in text[:400] and 'BANK' in text):
        return 'UBA'
    if 'ORABANK' in text:
        return 'ORABANK'
    if 'BANK OF AFRICA' in text:
        return 'BOA'
    if 'ARAB TUNISIAN BANK' in text:
        return 'ATB'
    if ('BSIC' in text or 'BANQUE SAHELO' in text or 'SN08SN111' in text_nospace):
        return 'BSIC'
    if ('BANQUE ISLAMIQUE DU SENEGAL' in text or 'ISLAMIQUE' in text and 'SENEGAL' in text):
        return 'BIS'
    if 'BNDE' in text or 'BANQUE NATIONALE POUR LE DEVELOPPEMENT' in text:
        return 'BNDE'
    return 'UNIVERSAL'


# ════════════════════════════════════════════════════════════════════════════
# PARSEURS BANCAIRES (identiques à la version Tkinter — non modifiés)
# ════════════════════════════════════════════════════════════════════════════

def parse_qonto(pages_words, pages_text):
    info = _extract_qonto_header(pages_text[0])
    year = int(info['period_start'].split('/')[2]) if info.get('period_start') else _year_from_text(pages_text[0])
    txns = []
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _qonto_date(row)
            if not date_str:
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 130 <= w['x0'] < 410).strip()
            amount = _qonto_amount(row)
            memo = ''
            j = i + 1
            while j < len(rows) and not _qonto_date(rows[j]):
                nl = ' '.join(w['text'] for w in rows[j] if 130 <= w['x0'] < 410).strip()
                na = _qonto_amount(rows[j])
                if na is not None and amount is None:
                    amount = na; memo = nl; j += 1; break
                elif na is None and nl:
                    memo = nl; j += 1; break
                else:
                    break
            i = j
            if amount is None or not label or label in ('Transactions', 'Date de valeur'):
                continue
            memo_clean = memo if memo.strip() not in ('', '-', '+') else ''
            name, memo_out = smart_label(label, [memo_clean] if memo_clean else [])
            txns.append(_make_txn(date_jjmm_to_ofx(date_str, year), amount, name, memo_out))
    return info, [t for t in txns if t is not None]

def _qonto_date(row):
    for w in row:
        if w['x0'] < 120 and re.match(r'^\d{2}/\d{2}$', w['text']):
            return w['text']
    return ''

def _qonto_amount(row):
    aw = [w for w in row if w['x0'] >= 400]
    if not aw: return None
    full = ' '.join(w['text'] for w in aw).replace('EUR','').replace('\xa0',' ').strip()
    m = re.search(r'([+\-])\s*([\d\s]+[.,]\d{2})', full)
    if m:
        sign = 1.0 if m.group(1)=='+' else -1.0
        try: return sign * float(m.group(2).replace(' ','').replace(',','.'))
        except: pass
    m2 = re.search(r'([\d\s]+[.,]\d{2})', full)
    if m2:
        sign = 1.0
        for w in aw:
            if w['text'] in ('+','-'): sign = 1.0 if w['text']=='+' else -1.0; break
            sm = re.match(r'^([+\-])([\d,.]+)$', w['text'])
            if sm: sign = 1.0 if sm.group(1)=='+' else -1.0; break
        try: return sign * float(m2.group(1).replace(' ','').replace(',','.'))
        except: pass
    return None

def _extract_qonto_header(text):
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    m = re.search(r'Du\s+(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})', text)
    if m: info['period_start'], info['period_end'] = m.group(1), m.group(2)
    bals = re.findall(r'Solde au \d{2}/\d{2}\s*[+\-]\s*([\d]+\.[\d]{2})\s*EUR', text)
    if len(bals) >= 1: info['balance_open']  = float(bals[0])
    if len(bals) >= 2: info['balance_close'] = float(bals[-1])
    return info

def parse_lcl(pages_words, pages_text):
    info = _extract_lcl_header(pages_text)
    year = _year_from_text(pages_text[0])
    txns = []

    # Labels à ignorer absolument (lignes de solde, totaux, en-têtes)
    SKIP_LABELS = {
        'DEBIT', 'CREDIT', 'VALEUR', 'DATE', 'LIBELLE', 'ANCIEN SOLDE',
        'SOLDE EN EUROS', 'TOTAUX', 'NOUVEAU SOLDE', 'SOLDE INITIAL',
        'SOLDE FINAL', 'TOTAL', 'REPORT', 'A REPORTER',
    }
    # Préfixes de labels de solde ou total (comparaison startswith)
    SKIP_PREFIXES = ('SOLDE', 'TOTAUX', 'TOTAL', 'ANCIEN', 'NOUVEAU')

    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _lcl_date(row)
            if not date_str:
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 70 <= w['x0'] < 360).strip()

            # Ignorer les lignes de solde/totaux même si elles ont une date
            label_up = label.upper().strip()
            if (label_up in SKIP_LABELS
                    or any(label_up.startswith(p) for p in SKIP_PREFIXES)
                    or not label_up):
                i += 1; continue

            debit_words = [w for w in row if 360 <= w['x0'] < 490
                           and not re.match(r'^\d{2}\.\d{2}(\.\d{2,4})?$', w['text'])]
            debit_amt  = _parse_col_amount(debit_words)
            credit_amt = _parse_col_amount([w for w in row if w['x0'] >= 490])

            memo = ''
            j = i + 1
            while j < len(rows) and not _lcl_date(rows[j]):
                nl = ' '.join(w['text'] for w in rows[j] if 70 <= w['x0'] < 360).strip()
                nl_up = nl.upper().strip()
                if nl and nl_up not in SKIP_LABELS and not any(nl_up.startswith(p) for p in SKIP_PREFIXES):
                    memo = (memo + ' ' + nl).strip()
                j += 1
            i = j

            date_ofx = date_jjmm_to_ofx(date_str, year)
            name, memo_out = smart_label(label, [memo] if memo else [])
            if debit_amt is not None:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo_out))
            elif credit_amt is not None:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo_out))
    return info, [t for t in txns if t is not None]

def _lcl_date(row):
    for w in row:
        if w['x0'] < 100 and re.match(r'^\d{2}\.\d{2}$', w['text']):
            return w['text']
    return ''

def _extract_lcl_header(pages_text):
    text = ' '.join(pages_text)
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)

    # Période : "du 03.10.2025 au 31.10.2025"
    m = re.search(r'du\s+(\d{2}\.\d{2}\.\d{4})\s+au\s+(\d{2}\.\d{2}\.\d{4})', text, re.IGNORECASE)
    if m:
        info['period_start'] = m.group(1).replace('.','/')
        info['period_end']   = m.group(2).replace('.','/')

    # Solde d'ouverture : "ANCIEN SOLDE  40 978,70" ou "ANCIEN SOLDE 40978,70"
    m_open = re.search(r'ANCIEN SOLDE\s+([\d\s]+[,\.]\d{2})', text)
    if m_open:
        v = parse_amount(m_open.group(1).replace(' ', ''))
        if v: info['balance_open'] = v

    # Solde de clôture : "SOLDE EN EUROS  46 862,54" (dernière ligne du relevé)
    for m_close in re.finditer(r'SOLDE EN EUROS\s+([\d\s]+[,\.]\d{2})', text):
        v = parse_amount(m_close.group(1).replace(' ', ''))
        if v: info['balance_close'] = v

    return info

def _ca_parse_zone(row, x_min, x_max):
    col = [w for w in row if x_min <= w['x0'] < x_max and re.match(r'^\d', w['text'])]
    if not col: return None
    last = col[-1]['text']
    if not re.match(r'^\d+,\d{2}$', last): return None
    if len(col) == 1: return parse_amount(last)
    prefix_tokens = [w['text'] for w in col[:-1]]
    if all(re.match(r'^\d+$', p) for p in prefix_tokens):
        try:
            return float(''.join(prefix_tokens) + last.replace(',', '.'))
        except ValueError:
            pass
    return None

def parse_ca(pages_words, pages_text):
    info = _extract_ca_header(pages_text)
    year = _year_from_text(pages_text[0])
    txns = []
    SKIP = {'Débit','Crédit','Date','Libellé','Total des opérations','Nouveau solde',
            'opé.','valeur','Libellé des opérations','Ancien solde débiteur','Nouveau solde débiteur'}
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _ca_date(row)
            if not date_str:
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 70 <= w['x0'] < 420).strip()
            debit_amt  = _ca_parse_zone(row, 415, 490)
            credit_amt = _ca_parse_zone(row, 490, 560)
            memo_parts = []
            j = i + 1
            while j < len(rows) and not _ca_date(rows[j]):
                nl = ' '.join(w['text'] for w in rows[j] if 70 <= w['x0'] < 420).strip()
                if nl and nl not in SKIP and len(nl) > 1:
                    memo_parts.append(nl)
                j += 1
            i = j
            if not label or label in SKIP: continue
            date_ofx = date_jjmm_to_ofx(date_str, year)
            name, memo = smart_label(label, memo_parts)
            if debit_amt is not None:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt is not None:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))
    return info, [t for t in txns if t is not None]

def _ca_date(row):
    for w in row:
        if w['x0'] < 50 and re.match(r'^\d{2}\.\d{2}$', w['text']):
            return w['text']
    return ''

def _extract_ca_header(pages_text):
    text = ' '.join(pages_text[:2])
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    mois_map = {'janvier':'01','février':'02','mars':'03','avril':'04','mai':'05','juin':'06',
                'juillet':'07','août':'08','septembre':'09','octobre':'10','novembre':'11','décembre':'12'}
    m = re.search(r'Date d.arrêté\s*:\s*(\d+)\s+(\w+)\s+(\d{4})', text)
    if m:
        mn = mois_map.get(m.group(2).lower(), '01')
        info['period_end']   = f"{m.group(1).zfill(2)}/{mn}/{m.group(3)}"
        info['period_start'] = f"01/{mn}/{m.group(3)}"
    return info

def parse_ce(pages_words, pages_text):
    info = _extract_ce_header(pages_text)
    txns = []
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _ce_date(row)
            if not date_str:
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 155 <= w['x0'] < 500).strip()
            amount = _parse_signed_amount([w for w in row if w['x0'] >= 500])
            memo = ''
            j = i + 1
            while j < len(rows) and not _ce_date(rows[j]):
                nl = ' '.join(w['text'] for w in rows[j] if 155 <= w['x0'] < 500).strip()
                if nl and len(nl) > 2:
                    memo = (memo + ' ' + nl).strip()
                j += 1
            i = j
            if not label or amount is None: continue
            skip_kw = {'DATE','VALEUR','MONTANT','OPERATIONS','SOLDE','TOTAL','DETAIL'}
            if any(s in label.upper() for s in skip_kw): continue
            date_ofx = date_full_to_ofx(date_str)
            name, memo_out = smart_label(label, [memo] if memo else [])
            txns.append(_make_txn(date_ofx, amount, name, memo_out))
    return info, [t for t in txns if t is not None]

def _ce_date(row):
    for w in row:
        if w['x0'] < 100 and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text']):
            return w['text']
    return ''

def _extract_ce_header(pages_text):
    text = ' '.join(pages_text[:2])
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    return info

def parse_bp(pages_words, pages_text):
    info = _extract_bp_header(pages_text)
    year = _year_from_text(pages_text[0])
    txns = []
    SKIP_KW = {'DATE','LIBELLE','REFERENCE','COMPTA','VALEUR','MONTANT','SOLDE','TOTAL','DETAIL','OPERATION'}
    for pw in pages_words:
        rows = group_words_by_row(pw)
        skip_from = None
        for idx, row in enumerate(rows):
            row_text = ' '.join(w['text'] for w in row).upper()
            if 'DETAIL DE VOS MOUVEMENTS SEPA' in row_text or 'DETAIL DE VOS PRELEVEMENTS SEPA' in row_text:
                skip_from = idx; break
        i = 0
        while i < len(rows):
            if skip_from is not None and i >= skip_from: break
            row = rows[i]
            date_str = _bp_date(row)
            if not date_str:
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 80 <= w['x0'] < 355).strip()
            amount = _bp_amount([w for w in row if w['x0'] >= 490])
            memo_parts = []
            j = i + 1
            while j < len(rows) and not _bp_date(rows[j]):
                if skip_from is not None and j >= skip_from: break
                nl = ' '.join(w['text'] for w in rows[j] if 80 <= w['x0'] < 355).strip()
                if nl and len(nl) > 2 and not re.match(r'^[\d\s.,€%=\-EUR]+$', nl):
                    memo_parts.append(nl)
                j += 1
            i = j
            if not label or amount is None: continue
            if any(s in label.upper() for s in SKIP_KW): continue
            date_ofx = date_jjmm_to_ofx(date_str, year)
            name, memo = smart_label(label, memo_parts)
            txns.append(_make_txn(date_ofx, amount, name, memo))
    return info, [t for t in txns if t is not None]

def _bp_date(row):
    for w in row:
        if w['x0'] < 80 and re.match(r'^\d{2}/\d{2}$', w['text']):
            return w['text']
    return ''

def _bp_amount(words):
    if not words: return None
    full = ' '.join(w['text'] for w in words).replace('€','').replace('\xa0',' ').strip()
    m = re.search(r'-\s*([\d\s]+[,.][\d]{2})', full)
    if m:
        try: return -abs(float(m.group(1).replace(' ','').replace(',','.')))
        except: pass
    m2 = re.search(r'\+\s*([\d\s]+[,.][\d]{2})', full)
    if m2:
        try: return abs(float(m2.group(1).replace(' ','').replace(',','.')))
        except: pass
    m3 = re.search(r'([\d\s]+[,.][\d]{2})', full)
    if m3:
        try:
            val = float(m3.group(1).replace(' ','').replace(',','.'))
            return val if val > 0 else None
        except: pass
    return None

def _extract_bp_header(pages_text):
    text = pages_text[0] if pages_text else ''
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    return info

def parse_cic(pages_words, pages_text):
    info = _extract_cic_header(pages_text)
    txns = []
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _cic_date(row)
            if not date_str:
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 140 <= w['x0'] < 430).strip()
            debit_amt  = _parse_col_amount([w for w in row if 420 <= w['x0'] < 500])
            credit_amt = _parse_col_amount([w for w in row if w['x0'] >= 500])
            memo = ''
            j = i + 1
            while j < len(rows) and not _cic_date(rows[j]):
                nl = ' '.join(w['text'] for w in rows[j] if 140 <= w['x0'] < 430).strip()
                if nl and len(nl) > 2 and not re.match(r'^[\d.,]+$', nl):
                    memo = (memo + ' ' + nl).strip()
                j += 1
            i = j
            if not label: continue
            skip_kw = {'DATE','DÉBIT','CRÉDIT','EUROS','SOLDE CREDITEUR','CREDIT INDUSTRIEL','TOTAL DES MOUVEMENTS'}
            if any(s in label.upper() for s in skip_kw): continue
            date_ofx = date_full_to_ofx(date_str)
            name, memo_out = smart_label(label, [memo] if memo else [])
            if debit_amt is not None:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo_out))
            elif credit_amt is not None:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo_out))
    return info, [t for t in txns if t is not None]

def _cic_date(row):
    for w in row:
        if w['x0'] < 100 and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text']):
            return w['text']
    return ''

def _extract_cic_header(pages_text):
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    for pt in reversed(pages_text):
        iban = extract_iban(pt)
        if iban:
            info['iban'] = iban; break
    return info

def parse_cgd(pages_words, pages_text):
    info = _extract_cgd_header(pages_text)
    year = _year_from_text(pages_text[0])
    txns = []
    SKIP = {'A REPORTER','REPORT','TOTAL','NOUVEAU','ANCIEN','SARL','CPT ORD'}
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            if not (len(row) >= 2
                    and re.match(r'^\d{2}$', row[0]['text']) and row[0]['x0'] < 50
                    and re.match(r'^\d{2}$', row[1]['text']) and row[1]['x0'] < 55):
                i += 1; continue
            dd, mm = row[0]['text'], row[1]['text']
            label = ' '.join(w['text'] for w in row if 70 <= w['x0'] < 310).strip()
            if not label or any(s in label.upper() for s in SKIP):
                i += 1; continue
            debit_amt  = _cgd_amount_in_zone(row, 395, 500)
            credit_amt = _cgd_amount_in_zone(row, 500, 570)
            memo_parts = []
            j = i + 1
            while j < len(rows):
                r2 = rows[j]
                if len(r2) >= 2 and re.match(r'^\d{2}$', r2[0]['text']) and r2[0]['x0'] < 50: break
                nl = ' '.join(w['text'] for w in r2 if 70 <= w['x0'] < 310).strip()
                if nl: memo_parts.append(nl)
                j += 1
            i = j
            date_ofx = f"{year}{mm.zfill(2)}{dd.zfill(2)}"
            name, memo = smart_label(label, memo_parts)
            if debit_amt:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))
    return info, [t for t in txns if t is not None]

def _cgd_amount_in_zone(row, x_min, x_max):
    col = [w for w in row if x_min <= w['x0'] < x_max and re.match(r'^\d', w['text'])]
    if not col: return None
    return parse_amount(col[-1]['text'])

def _extract_cgd_header(pages_text):
    text = ' '.join(pages_text[:2])
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    return info

def parse_lbp(pages_words, pages_text):
    info = _extract_lbp_header(pages_text)
    year = _year_from_text(pages_text[0])
    txns = []
    SKIP = {'TOTAL DES','NOUVEAU SOLDE','ANCIEN SOLDE','VOS OPERATIONS','DATE OPERATION','SITUATION DU','PAGE'}
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            if not (row[0]['x0'] < 60 and re.match(r'^\d{2}/\d{2}$', row[0]['text'])):
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 85 <= w['x0'] < 430).strip()
            label = re.sub(r'\(cid:\d+\)', '', label).strip()
            if not label or any(s in label.upper() for s in SKIP):
                i += 1; continue
            debit_amt  = _lbp_amount_in_zone(row, 430, 500)
            credit_amt = _lbp_amount_in_zone(row, 500, 560)
            j = i + 1
            while j < len(rows):
                r2 = rows[j]
                if r2[0]['x0'] < 60 and re.match(r'^\d{2}/\d{2}$', r2[0]['text']): break
                j += 1
            i = j
            date_ofx = f"{year}{row[0]['text'][3:5]}{row[0]['text'][:2]}"
            name, memo = smart_label(label, [])
            if debit_amt:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))
    return info, [t for t in txns if t is not None]

def _lbp_amount_in_zone(row, x_min, x_max):
    col = [w for w in row if x_min <= w['x0'] < x_max and re.match(r'^\d', w['text'])]
    if not col: return None
    last = col[-1]['text']
    if not re.match(r'^\d+,\d{2}$', last): return None
    if len(col) == 1: return parse_amount(last)
    prefix_tokens = [w['text'] for w in col[:-1]]
    if all(re.match(r'^\d+$', p) for p in prefix_tokens):
        try: return float(''.join(prefix_tokens) + last.replace(',', '.'))
        except: pass
    return parse_amount(last)

def _extract_lbp_header(pages_text):
    text = re.sub(r'\(cid:\d+\)', ' ', ' '.join(pages_text[:2]))
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    return info

def parse_sg(pages_words, pages_text):
    info = _extract_sg_header(pages_text)
    txns = []
    SKIP = {'TOTAUX DES','NOUVEAU SOLDE','SOLDE PRECEDENT','PROGRAMME DE','RAPPEL DES','MONTANT CUMULE'}
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            if not (row[0]['x0'] < 45 and re.match(r'^\d{2}/\d{2}/\d{4}$', row[0]['text'])):
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 120 <= w['x0'] < 430).strip()
            if not label or any(s in label.upper() for s in SKIP):
                i += 1; continue
            debit_amt  = _sg_amount_in_zone(row, 430, 510)
            credit_amt = _sg_amount_in_zone(row, 510, 570)
            memo_parts = []
            j = i + 1
            while j < len(rows):
                r2 = rows[j]
                if r2[0]['x0'] < 45 and re.match(r'^\d{2}/\d{2}/\d{4}$', r2[0]['text']): break
                nl = ' '.join(w['text'] for w in r2 if 120 <= w['x0'] < 430).strip()
                if nl and not any(s in nl.upper() for s in ('TOTAUX','NOUVEAU','PROGRAMME','RAPPEL')):
                    memo_parts.append(nl)
                j += 1
            i = j
            date_ofx = date_full_to_ofx(row[0]['text'])
            name, memo = smart_label(label, memo_parts)
            if debit_amt:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))
    return info, [t for t in txns if t is not None]

def _uba_join_amount(words):
    """
    Reconstitue un montant XOF/EUR fragmente en plusieurs tokens pdfplumber.
    Gere les montants avec decimales : ['41', '295,00'] → 41295.0
    ET les montants entiers XOF     : ['15', '400']     → 15400.0
    ET les grands entiers           : ['2', '316', '032,00'] → 2316032.0
    Strategie :
      1. Cherche d'abord un pattern avec decimales (ex: '15 400,00')
      2. Sinon, si tous les tokens sont numeriques purs, les concatene
         en verifiant que c'est coherent (pas un solde, pas une date)
    """
    if not words:
        return None
    full = ' '.join(w['text'] for w in words).replace('\xa0', ' ').strip()
    if not full or full in ('.', ','):
        return None

    # --- Cas 0 : format BIS iMAL (virgule séparateur milliers, pas de décimales) ---
    # Ex: "2,362,500"  "660,000"  "1,350,412"
    m0 = re.match(r'^(\d{1,3}(?:,\d{3})+)$', full.replace(' ', ''))
    if m0:
        try:
            val = float(m0.group(1).replace(',', ''))
            if val > 0:
                return val
        except ValueError:
            pass

    # --- Cas 1 : montant avec decimales (EUR ou XOF) ---
    # Ex: "15 400,00"  "2 316 032,00"  "41,95"
    m = re.search(r'(\d[\d\s]*[\.,]\d{2})\b', full)
    if m:
        raw = m.group(1).strip()
        normalized = re.sub(r'\s+', '', raw).replace(',', '.')
        try:
            val = float(normalized)
            if val > 0:
                return val
        except ValueError:
            pass

    # --- Cas 2 : montant entier XOF (pas de decimales) ---
    # Tous les tokens doivent etre purement numeriques
    # Ex: ['15', '400']  ['369', '630']  ['5', '850']
    texts = [w['text'] for w in words]
    if all(re.match(r'^\d+$', t) for t in texts) and len(texts) >= 2:
        # Verifier que la concatenation donne un nombre raisonnable
        # Le 2e token et suivants doivent avoir exactement 3 chiffres
        # (separateur milliers) pour eviter de concatener une date valeur
        valid = True
        for t in texts[1:]:
            if len(t) != 3:
                valid = False
                break
        if valid:
            try:
                val = float(''.join(texts))
                if val > 0:
                    return val
            except ValueError:
                pass
    elif len(texts) == 1 and re.match(r'^\d+$', texts[0]):
        # Montant entier sur un seul token
        try:
            val = float(texts[0])
            if val > 0:
                return val
        except ValueError:
            pass

    return None

def _sg_amount_in_zone(row, x_min, x_max):
    col = [w for w in row if x_min <= w['x0'] < x_max]
    if not col: return None
    # Enlever le '*' (opérations exonérées de commission) avant de parser
    col_clean = [w for w in col if w['text'].strip() not in ('*', '')]
    if not col_clean: return None

    # Reconstituer la chaîne complète et tenter parse_amount directement
    # (gère correctement "1.082,92", "29.117,17", etc.)
    full = ' '.join(w['text'] for w in col_clean).replace('*', '').replace('\xa0', ' ').strip()
    # Format français groupé par 3 avec point comme séparateur de milliers : 1.082,92
    m_fr = re.search(r'(\d{1,3}(?:\.\d{3})+,\d{2})', full)
    if m_fr:
        v = parse_amount(m_fr.group(1))
        if v is not None and v > 0:
            return v
    # Format simple : 100,75  ou  312,48
    m_simple = re.search(r'(\d+,\d{2})', full)
    if m_simple:
        v = parse_amount(m_simple.group(1))
        if v is not None and v > 0:
            return v
    # Fallback : montants fragmentés en plusieurs tokens (ex: "5 850" → 5850)
    v = _uba_join_amount(col_clean)
    if v is not None:
        return v
    return None

def _extract_sg_header(pages_text):
    text = ' '.join(pages_text[:2])
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}

    # ── IBAN (présent sur certains relevés SG) ────────────────────────────────
    info['iban'] = extract_iban(text)

    # ── RIB SG : "n° 30003 03320 00020641644 69" ─────────────────────────────
    # Format : n° BBBBB GGGGG CCCCCCCCCCC KK  (Banque 5 + Guichet 5 + Compte 11 + Clé 2)
    # pdfplumber peut lire avec espaces OU en un seul token compact "n°30003033200002064164469"
    if not info['iban']:
        # Cas 1 : avec espaces
        rib_m = re.search(
            r'n[°o]\s*(\d{5})\s+(\d{5})\s+(\d{10,12})\s+(\d{2})\b', text)
        # Cas 2 : compact sans espaces (23 chiffres : 5+5+11+2)
        if not rib_m:
            rib_m = re.search(r'n[°o]\s*(\d{5})(\d{5})(\d{11})(\d{2})\b', text)
        if rib_m:
            banque  = rib_m.group(1)
            guichet = rib_m.group(2)
            compte  = rib_m.group(3)
            cle     = rib_m.group(4)
            info['_rib_bank']    = banque
            info['_rib_agency']  = guichet
            info['_rib_account'] = compte
            info['_rib_key']     = cle
            info['iban'] = f"{banque} {guichet} {compte} {cle}"

    # ── Période : "du 01/03/2026 au 31/03/2026" ──────────────────────────────
    # Tolérance aux espaces multiples et aux retours à la ligne (layout multi-col SG)
    m = re.search(
        r'du\s+(\d{2}/\d{2}/\d{4})\s+au\s+(\d{2}/\d{2}/\d{4})',
        text, re.IGNORECASE
    )
    if m:
        info['period_start'] = m.group(1)
        info['period_end']   = m.group(2)

    # ── Solde de clôture : "NOUVEAU SOLDE AU 31/03/2026 + 85.536,72" ─────────
    m_bal = re.search(
        r'NOUVEAU SOLDE\s+AU\s+\d{2}/\d{2}/\d{4}\s+[+\-]?\s*([\d\s]+[,\.]\d{2})',
        text, re.IGNORECASE
    )
    if m_bal:
        v = parse_amount(m_bal.group(1).replace(' ', '').replace('\xa0', ''))
        if v:
            info['balance_close'] = v

    return info

def parse_bnp(pages_words, pages_text):
    info = _extract_bnp_header(pages_text)
    year = _year_from_text(' '.join(pages_text[:2]))
    txns = []
    SKIP = {'DATE','LIBELLE','VALEUR','DEBIT','CREDIT','EUROS','SOLDE','TOTAL','OPERATIONS',
            'ANCIEN SOLDE','NOUVEAU SOLDE','VIREMENT RECU','RELEVE DE COMPTE'}
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _bnp_date(row)
            if not date_str:
                i += 1; continue
            label = ' '.join(w['text'] for w in row if 85 <= w['x0'] < 430).strip()
            if not label or any(s in label.upper() for s in SKIP):
                i += 1; continue
            debit_amt  = _parse_col_amount([w for w in row if 480 <= w['x0'] < 560])
            credit_amt = _parse_col_amount([w for w in row if w['x0'] >= 560])
            memo_parts = []
            j = i + 1
            while j < len(rows) and not _bnp_date(rows[j]):
                nl = ' '.join(w['text'] for w in rows[j] if 85 <= w['x0'] < 430).strip()
                if nl and len(nl) > 2:
                    memo_parts.append(nl)
                j += 1
            i = j
            date_ofx = _bnp_date_to_ofx(date_str, year)
            name, memo = smart_label(label, memo_parts)
            if debit_amt is not None:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt is not None:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))
    return info, [t for t in txns if t is not None]

def _bnp_date(row):
    for w in row:
        if w['x0'] < 80:
            if re.match(r'^\d{2}/\d{2}/\d{2}$', w['text']): return w['text']
            if re.match(r'^\d{2}/\d{2}/\d{4}$', w['text']): return w['text']
    return ''

def _bnp_date_to_ofx(date_str, year_hint):
    parts = date_str.split('/')
    if len(parts) == 3:
        dd, mm, yy = parts[0].zfill(2), parts[1].zfill(2), parts[2]
        if len(yy) == 2:
            full_year = (2000 + int(yy)) if int(yy) <= 30 else (1900 + int(yy))
        else:
            full_year = int(yy)
        return f"{full_year}{mm}{dd}"
    return str(year_hint) + '0101'

def _extract_bnp_header(pages_text):
    text = ' '.join(pages_text[:2])
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    return info

def parse_mypos(pages_words, pages_text):
    info = _extract_mypos_header(pages_text)
    txns = []
    full_text = '\n'.join(pages_text)
    lines = [l.strip() for l in full_text.splitlines()]
    txn_re = re.compile(
        r'^(\d{2}\.\d{2}\.\d{4})\s+\d{2}:\d{2}\s+'
        r'(System Fee|myPOS Payment|Glass Payment|Outgoing bank transfer|POS Payment|Mobile)\s*'
        r'.*?1\.0000\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s*$'
    )
    for idx, line in enumerate(lines):
        m = txn_re.match(line)
        if not m: continue
        date_raw = m.group(1)
        txn_type = m.group(2).strip()
        try:
            debit_val  = float(m.group(3).replace(',', ''))
            credit_val = float(m.group(4).replace(',', ''))
        except ValueError:
            continue
        date_ofx = f"{date_raw[6:10]}{date_raw[3:5]}{date_raw[0:2]}"
        description = ''
        for back in (1, 2):
            if idx >= back:
                prev = lines[idx - back].strip()
                if prev and not re.match(r'^\d{2}\.\d{2}\.\d{4}', prev):
                    description = prev; break
        if txn_type == 'System Fee':
            name, memo = 'myPOS Fee', description
        else:
            name, memo = description or txn_type, description
        amount = -debit_val if debit_val > 0 else (credit_val if credit_val > 0 else None)
        if amount is None: continue
        txns.append(_make_txn(date_ofx, amount, name[:64], memo[:128]))
    return info, [t for t in txns if t is not None]

def _extract_mypos_header(pages_text):
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    text = pages_text[0] if pages_text else ''
    m = re.search(r'IBAN\s*:?\s*(IE\d{2}[A-Z0-9]+)', text)
    if m: info['iban'] = m.group(1).replace(' ','')
    m2 = re.search(r'Monthly statement\s*[-–]\s*(\d{2})\.(\d{4})', text, re.IGNORECASE)
    if m2:
        import calendar
        month, year = m2.group(1), m2.group(2)
        last_day = calendar.monthrange(int(year), int(month))[1]
        info['period_start'] = f"01/{month}/{year}"
        info['period_end']   = f"{last_day:02d}/{month}/{year}"
    return info

def parse_shine(pages_words, pages_text):
    info = _extract_shine_header(pages_text)
    txns = []
    SKIP = {'DATE','TYPE','OPÉRATION','OPERATION','DÉBIT','DEBIT','CRÉDIT','CREDIT',
            '(EURO)','SOLDE','TOTAL','NOUVEAU','COMMISSIONS','MOUVEMENTS','PAGE','LES','RELEVÉ'}
    for pw in pages_words:
        rows = group_words_by_row(pw)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _shine_date(row)
            if not date_str:
                i += 1; continue
            txn_type = ' '.join(w['text'] for w in row if 95 <= w['x0'] < 160).strip()
            label    = ' '.join(w['text'] for w in row if 160 <= w['x0'] < 453).strip()
            debit_amt  = _parse_col_amount([w for w in row if 453 <= w['x0'] < 513])
            credit_amt = _parse_col_amount([w for w in row if w['x0'] >= 513])
            memo_parts = []
            j = i + 1
            while j < len(rows) and not _shine_date(rows[j]):
                nl = ' '.join(w['text'] for w in rows[j] if 95 <= w['x0'] < 453).strip()
                if nl and len(nl) > 2:
                    memo_parts.append(nl)
                j += 1
            i = j
            full_label = (txn_type + ' ' + label).strip() if txn_type else label
            if any(s in full_label.upper() for s in SKIP) or len(full_label) < 2: continue
            date_ofx = date_full_to_ofx(date_str)
            name, memo = smart_label(full_label, memo_parts)
            if debit_amt is not None:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt is not None:
                txns.append(_make_txn(date_ofx,  credit_amt, name, memo))
    return info, [t for t in txns if t is not None]

def _shine_date(row):
    for w in row:
        if w['x0'] < 60 and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text']):
            return w['text']
    return ''

def _extract_shine_header(pages_text):
    text = ' '.join(pages_text[:3])
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    info['iban'] = extract_iban(text)
    m = re.search(r'De\s+(\d{2}/\d{2}/\d{4})\s+[àa]\s+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
    if m:
        info['period_start'] = m.group(1); info['period_end'] = m.group(2)
    return info


# ════════════════════════════════════════════════════════════════════════════
# PARSEUR UNIVERSEL
# ════════════════════════════════════════════════════════════════════════════

_COL_SYNONYMS = {
    'date':   ['date','date opé','date opé.','date opération','date val','date valeur','date comptable','valeur','jour','date op'],
    'label':  ['libellé','libelle','opération','operation','description','motif','désignation','nature','détail','detail','mouvement','intitulé','label','wording','particulars','narration'],
    'debit':  ['débit','debit','débit (euro)','debit (euro)','sorties','sortie','retrait','retraits','paiements','débit fcfa','débit xof','withdrawals','withdrawal','payments','dr','déb','deb'],
    'credit': ['crédit','credit','crédit (euro)','credit (euro)','entrées','entrée','versement','versements','encaissements','crédit fcfa','crédit xof','deposits','deposit','receipts','cr','cré','cred'],
    'amount': ['montant','amount','somme','mouvement','débit/crédit','debit/credit','montant net','net'],
    'balance':['solde','balance','solde après','running balance'],
}

_DATE_PATTERNS = [
    (r'^(\d{2})[/\-\.](\d{2})[/\-\.](\d{4})$', 'dmy4'),
    (r'^(\d{4})[/\-\.](\d{2})[/\-\.](\d{2})$', 'ymd4'),
    (r'^(\d{2})[/\-\.](\d{2})[/\-\.](\d{2})$', 'dmy2'),
    (r'^(\d{2})[/\-\.](\d{2})$',                'dm'),
    (r'^(\d{8})$',                               'Ymd8'),
]

def _match_col(cell_text, col_type):
    if not cell_text: return False
    t = str(cell_text).strip().lower()
    for syn in _COL_SYNONYMS[col_type]:
        if t == syn or t.startswith(syn + ' ') or t.startswith(syn + '('): return True
    return False

def _detect_header_row(table):
    for row_idx, row in enumerate(table[:20]):
        col_map = {}
        for col_idx, cell in enumerate(row):
            if not cell: continue
            for ctype in ('date','label','debit','credit','amount','balance'):
                if ctype not in col_map and _match_col(str(cell), ctype):
                    col_map[ctype] = col_idx
        has_date  = 'date' in col_map
        has_money = ('debit' in col_map and 'credit' in col_map) or 'amount' in col_map
        if has_date and has_money:
            return row_idx, col_map
    return None, {}

def _parse_date_universal(raw, year_hint=None):
    if not raw: return None
    raw = str(raw).strip()
    raw = re.sub(r'^[A-Za-zÀ-ÿ]+\.?\s*', '', raw).strip()
    raw = raw.split('\n')[0].strip()
    for pattern, fmt in _DATE_PATTERNS:
        m = re.match(pattern, raw)
        if not m: continue
        if fmt == 'dmy4': return f"{m.group(3)}{m.group(2).zfill(2)}{m.group(1).zfill(2)}"
        elif fmt == 'ymd4': return f"{m.group(1)}{m.group(2).zfill(2)}{m.group(3).zfill(2)}"
        elif fmt == 'dmy2':
            yy = int(m.group(3))
            return f"{2000+yy if yy<=30 else 1900+yy}{m.group(2).zfill(2)}{m.group(1).zfill(2)}"
        elif fmt == 'dm':
            yr = str(year_hint) if year_hint else str(datetime.now().year)
            return f"{yr}{m.group(2).zfill(2)}{m.group(1).zfill(2)}"
        elif fmt == 'Ymd8':
            s = m.group(0); return f"{s[:4]}{s[4:6]}{s[6:8]}"
    return None

def _parse_amount_cell(cell_text):
    """
    Parse un montant dans une cellule de tableau.
    Gère : EUR (1 234,56), XOF entiers (1 234 567), formats anglais (1,234.56),
    négatifs entre parenthèses, signes +/-.
    """
    if not cell_text: return None
    s = str(cell_text).strip().replace('\xa0',' ').replace('\u202f',' ').replace('\n',' ').strip()
    s = re.sub(r'[€$£FCFAXOF]','',s,flags=re.IGNORECASE).strip().replace('*','').strip()
    if not s or s in ('.', ',', '-', '—', '–', ''): return None
    negative = False
    if s.startswith('(') and s.endswith(')'): s = s[1:-1].strip(); negative = True
    if s.startswith('-'): negative = True; s = s[1:].strip()
    elif s.startswith('+'): s = s[1:].strip()
    s_nospace = s.replace(' ','')
    # Format EUR : 1.234,56 ou 1,234.56
    m = re.match(r'^(\d{1,3}(?:[.,]\d{3})+)[,.](\d{2})$', s_nospace)
    if m:
        integer_part = re.sub(r'[,.]','',m.group(1))
        val = float(f"{integer_part}.{m.group(2)}")
        return -val if negative else val
    # Format simple avec décimales
    m2 = re.match(r'^(\d+)[,.](\d{1,2})$', s_nospace)
    if m2:
        val = float(f"{m2.group(1)}.{m2.group(2)}")
        return -val if negative else val
    # ── XOF / entiers purs (séparateurs d'espaces ou points de milliers) ───────
    # Ex : "1 234 567" ou "1.234.567" (sans décimale)
    m3_spaces = re.match(r'^(\d{1,3}(?:\s\d{3})+)$', s)
    if m3_spaces:
        val = float(s.replace(' ', ''))
        return -val if negative else val
    m3_dots = re.match(r'^(\d{1,3}(?:\.\d{3})+)$', s_nospace)
    if m3_dots:
        val = float(s_nospace.replace('.', ''))
        return -val if negative else val
    # Entier brut
    m3 = re.match(r'^\d[\d\s]*\d$|^\d$', s)
    if m3:
        val = float(s.replace(' ',''))
        return -val if negative else val
    return None

def _extract_universal_header(pages_text):
    text = ' '.join(pages_text[:3])
    info = {'iban':'','period_start':'','period_end':'','balance_open':0.0,'balance_close':0.0}
    # Utilise extract_iban() centralisé pour cohérence avec tous les autres parsers
    info['iban'] = extract_iban(text)
    m1 = re.search(
        r'(?:du|from|de|period[e]?\s*:?)\s*(\d{1,2}[/\-.]\d{2}[/\-.]\d{2,4})'
        r'\s*(?:au|to|[àa]|\-)\s*(\d{1,2}[/\-.]\d{2}[/\-.]\d{2,4})',
        text, re.IGNORECASE)
    if m1:
        info['period_start'] = m1.group(1).replace('-','/').replace('.','/') 
        info['period_end']   = m1.group(2).replace('-','/').replace('.','/')
    return info

def _universal_parse_path(pdf_path, pages_text):
    """
    Parseur universel fiabilisé :
    - Stratégie 1 : détection tableau pdfplumber (3 configurations)
    - Stratégie 2 : fallback extraction texte ligne par ligne
    Supporte EUR (décimales) et XOF (entiers fragmentés).
    """
    info = _extract_universal_header(pages_text)
    year_hint = _year_from_text(' '.join(pages_text[:2]))
    txns = []
    SKIP_LABELS = {'TOTAL','TOTAUX','SOLDE','SOUS-TOTAL','REPORT','A REPORTER',
                   'NOUVEAU SOLDE','ANCIEN SOLDE','SOLDE INITIAL','SOLDE FINAL'}
    TABLE_SETTINGS_LIST = [
        {"vertical_strategy":"text","horizontal_strategy":"text","snap_tolerance":4,"join_tolerance":4},
        {"vertical_strategy":"lines","horizontal_strategy":"lines","snap_tolerance":3},
        {"vertical_strategy":"lines","horizontal_strategy":"text","snap_tolerance":4},
        {"vertical_strategy":"text","horizontal_strategy":"lines","snap_tolerance":5,"join_tolerance":5},
    ]

    # ── Stratégie 1 : extraction via tableau pdfplumber ───────────────────────
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                table = None
                for settings in TABLE_SETTINGS_LIST:
                    t = page.extract_table(settings)
                    if t and len(t) >= 3:
                        table = t; break
                if not table:
                    continue
                table_clean = [[str(c).replace('\n',' ').strip() if c else '' for c in row] for row in table]
                header_idx, col_map = _detect_header_row(table_clean)
                if header_idx is None:
                    continue
                for row in table_clean[header_idx + 1:]:
                    if not any(row):
                        continue
                    date_col = col_map.get('date')
                    if date_col is None or date_col >= len(row):
                        continue
                    date_ofx = _parse_date_universal(row[date_col], year_hint)
                    if not date_ofx:
                        continue
                    label_col = col_map.get('label')
                    label = row[label_col].strip() if (label_col is not None and label_col < len(row)) else row[date_col]
                    label_up = label.upper().strip()
                    if not label or len(label) < 2:
                        continue
                    if any(skip in label_up for skip in SKIP_LABELS):
                        continue
                    if re.match(r'^[\d\s.,\-]+$', label):
                        continue
                    amount = None
                    if 'debit' in col_map and 'credit' in col_map:
                        d_col, c_col = col_map['debit'], col_map['credit']
                        dv = _parse_amount_cell(row[d_col] if d_col < len(row) else '')
                        cv = _parse_amount_cell(row[c_col] if c_col < len(row) else '')
                        if dv and dv > 0: amount = -dv
                        elif cv and cv > 0: amount = cv
                    elif 'amount' in col_map:
                        a_col = col_map['amount']
                        amount = _parse_amount_cell(row[a_col] if a_col < len(row) else '')
                    if amount is None or amount == 0.0:
                        continue
                    name, memo = smart_label(label, [])
                    txn = _make_txn(date_ofx, amount, name, memo)
                    if txn:
                        txns.append(txn)
    except Exception as exc:
        logger.warning("_universal_parse_path tableau échoué : %s", exc)

    if txns:
        return info, [t for t in txns if t is not None]

    # ── Stratégie 2 : fallback texte ligne par ligne ──────────────────────────
    # Utile quand pdfplumber ne détecte pas de tableau (PDF scanné OCR ou layout libre)
    logger.info("_universal_parse_path : fallback texte brut ligne par ligne")
    CREDIT_KW = ('VERSEMENT','VIREMENT RECU','VRT RECU','REMISE','CREDIT','AVOIR',
                 'RETOUR','REMBOURSEMENT','SWIFT ENTRANT','DECAISSEMENT','DEBLOCAGE',
                 'TRF RECU','V/V FAC')
    DEBIT_KW  = ('RETRAIT','CHEQUE','CHQ','FRAIS','COMMISSION','AGIOS','PRELEVEMENT',
                 'PRLV','ABONNEMENT','IMPAYE','COTISATION','VIREMENT EMIS','VRT EMIS',
                 'FACTURATION','TENUE DE COMPTE','PACKAGE','REDEVANCE','IBE')

    full_text = '\n'.join(pages_text).replace('\xa0', ' ').replace('\u202f', ' ')
    lines = full_text.split('\n')
    prev_solde = None

    for line in lines:
        line = line.strip()
        m_date = re.match(r'^(\d{2}[/\-.]\d{2}[/\-.]\d{2,4})', line)
        if not m_date:
            continue
        date_str = m_date.group(1)
        date_ofx = _parse_date_universal(date_str, year_hint)
        if not date_ofx:
            continue
        line_up = line.upper()
        if any(kw in line_up for kw in ('SOLDE','TOTAL','LIBELLÉ','LIBELLE',
                                          'DÉBIT','DEBIT','CRÉDIT','CREDIT',
                                          'EXTRAIT','PAGE','PÉRIODE')):
            continue
        # Extraire tous les montants (XOF entiers >= 100, ou EUR avec décimales)
        raw_amounts = re.findall(
            r'\b(\d{1,3}(?:\s\d{3})*(?:[.,]\d{2})?)\b', line)
        amounts_vals = []
        for a in raw_amounts:
            v = _parse_amount_cell(a)
            if v is not None and v >= 100:
                amounts_vals.append(v)
        if not amounts_vals:
            continue
        # Libellé : enlever date(s) de début + montants de fin
        label_part = re.sub(r'^\d{2}[/\-.]\d{2}[/\-.]\d{2,4}\s*', '', line)
        label_part = re.sub(r'^\d{2}[/\-.]\d{2}[/\-.]\d{2,4}\s*', '', label_part)
        label_part = re.sub(r'[\d\s,.]+$', '', label_part).strip()
        label_part = clean_label(label_part)
        if not label_part or len(label_part) < 3:
            continue
        label_up2 = label_part.upper()
        if any(skip in label_up2 for skip in SKIP_LABELS):
            continue
        # Déduire débit/crédit
        is_credit = None
        amt = amounts_vals[-1] if len(amounts_vals) == 1 else amounts_vals[-2]
        if len(amounts_vals) >= 2:
            solde_courant = amounts_vals[-1]
            if prev_solde is not None:
                diff = solde_courant - prev_solde
                if diff > 10: is_credit = True
                elif diff < -10: is_credit = False
            prev_solde = solde_courant
        if is_credit is None:
            if any(k in label_up2 for k in CREDIT_KW): is_credit = True
            elif any(k in label_up2 for k in DEBIT_KW): is_credit = False
            else: is_credit = False  # défaut débit
        signed_amt = amt if is_credit else -amt
        txn = _make_txn(date_ofx, signed_amt, label_part)
        if txn:
            txns.append(txn)

    return info, [t for t in txns if t is not None]

# ════════════════════════════════════════════════════════════════════════════
# PARSEURS AFRICAINS DÉDIÉS
# ════════════════════════════════════════════════════════════════════════════

def _afr_header(pages_text):
    """Header commun pour les banques africaines."""
    text = ' '.join(pages_text[:3])
    info = {'iban': '', 'period_start': '', 'period_end': '',
            'balance_open': 0.0, 'balance_close': 0.0}

    # ── 1. RIB explicitement labellisé : Code Banque / Agence / Compte / Clé ──
    # Ex tableau : "Code Banque  Agence   Compte         Clé RIB"
    #              "SN213        01001    02341624101    33"
    rib_table = re.search(
        r'(?:Code\s*Banque|Banque)\s*[:\|]?\s*([A-Z]{0,2}\d{3,5})'
        r'.{0,40?}(?:Agence|Guichet)\s*[:\|]?\s*(\d{4,6})'
        r'.{0,40?}(?:N[°o]?\s*(?:de\s*)?[Cc]ompte|Compte)\s*[:\|]?\s*(\d{8,14})'
        r'.{0,40?}(?:Cl[eé]\s*(?:RIB)?|RIB)\s*[:\|]?\s*(\d{2})\b',
        text, re.IGNORECASE | re.DOTALL
    )
    if rib_table:
        info['_rib_bank']    = rib_table.group(1)
        info['_rib_agency']  = rib_table.group(2)
        info['_rib_account'] = rib_table.group(3)
        info['_rib_key']     = rib_table.group(4)

    # ── 2. RIB compact sur une ligne : "SN213 01001 02341624101 33" ───────────
    _SEP = r'[\s\t\-]+'
    if not info.get('_rib_bank'):
        rib_match = re.search(
            r'\b([A-Z]{0,2}\d{3,5})' + _SEP +
            r'(\d{4,6})'             + _SEP +
            r'(\d{8,14})'            + _SEP +
            r'(\d{2})\b',
            text
        )
        if rib_match:
            info['_rib_bank']    = rib_match.group(1)
            info['_rib_agency']  = rib_match.group(2)
            info['_rib_account'] = rib_match.group(3)
            info['_rib_key']     = rib_match.group(4)

    # ── 2b. RIB format tiret BSIC/ECOBANK : "01001-00100029193-76" ─────────────
    if not info.get('_rib_bank'):
        rib_tiret = re.search(
            r'(?:Num[\xe9e]ro\s+de\s+compte|N[o\xb0]\s*compte)\s*[:\-]?\s*'
            r'(\d{5})-(\d{8,14})-(\d{2})',
            text, re.IGNORECASE
        )
        if rib_tiret:
            info['_rib_agency']  = rib_tiret.group(1)
            info['_rib_account'] = rib_tiret.group(2)
            info['_rib_key']     = rib_tiret.group(3)

    # ── 3. Extraction IBAN centralisée ───────────────────────────────────────
    if not info['iban']:
        info['iban'] = extract_iban(text)

    # ── 4. Si IBAN trouvé mais pas de code banque, dériver depuis l'IBAN ──────
    # Ne pas ecraser _rib_account si deja extrait par regex tiret (plus fiable)
    if info['iban'] and not info.get('_rib_bank'):
        try:
            b, ag, ac = iban_to_rib(info['iban'])
            if b and b != '00000':
                info['_rib_bank']   = b
                info['_rib_agency'] = info.get('_rib_agency') or ag
                # Ne pas ecraser le compte si deja capture par la regex tiret
                if not info.get('_rib_account'):
                    info['_rib_account'] = ac
                info['_rib_key']    = info.get('_rib_key') or ''
        except Exception:
            pass

    # ── 5. Numéro de compte brut (dernier recours) ───────────────────────────
    if not info['iban']:
        m4 = re.search(
            r'(?:N[°o°]\.?\s*(?:de\s*)?compte|Compte|COMPTE)\s*[:\-]?\s*'
            r'([\d]{5,20}(?:[\s\-]?\d{1,6})*)',
            text, re.IGNORECASE
        )
        if m4:
            info['iban'] = re.sub(r'[\s\-]', '', m4.group(1))

    # ── 6. Période ───────────────────────────────────────────────────────────
    m5 = re.search(r'(?:du|[Pp]ériode du?|[Pp]our la p[ée]riode du?|[Dd]u)\s+'
                   r'(\d{1,2}[/\-\.]\d{2}[/\-\.]\d{2,4})'
                   r'\s*(?:au|[àa]|[Aa]u)\s*'
                   r'(\d{1,2}[/\-\.]\d{2}[/\-\.]\d{2,4})',
                   text, re.IGNORECASE)
    if m5:
        info['period_start'] = m5.group(1).replace('-', '/').replace('.', '/')
        info['period_end']   = m5.group(2).replace('-', '/').replace('.', '/')
    return info


# ── BSIC : format « Date Op | Date Val | Libellé | Débit | Crédit | Solde »
# Positions mesurées sur PDF réel :
#   Date opération : x0 ≈ 32    (dd/mm/yyyy)
#   Date valeur    : x0 ≈ 80    (dd/mm/yyyy) — à exclure du libellé
#   Libellé        : x0 ≈ 128-320
#   Débit          : x0 ≈ 355-410
#   Crédit         : x0 ≈ 440-500
#   Solde          : x0 ≈ 510+  (ignoré)
def parse_bsic(pages_words, pages_text, _pdf_path=''):
    info = _afr_header(pages_text)
    year = _year_from_text(' '.join(pages_text[:2]))
    txns = []
    SKIP = {'TOTAL','SOLDE','A REPORTER','REPORT','DATE','VALEUR','LIBELLÉ','LIBELLE',
            'DÉBIT','DEBIT','CRÉDIT','CREDIT','EXTRAIT','PÉRIODE','CODE','NOM',
            'PAGE SUR'}  # "PAGE" et "SUR" uniquement en mots entiers via regex ci-dessous

    def _is_skip_label(lbl):
        """Retourne True si le libellé est une ligne d'en-tête/pied à ignorer."""
        lbl_up = lbl.upper().strip()
        # Mots-clés exacts (libellé = entièrement ce mot)
        if lbl_up in SKIP:
            return True
        # Mots-clés en début de libellé (solde, total…)
        for kw in ('SOLDE', 'TOTAL', 'A REPORTER', 'REPORT', 'DÉBIT', 'DEBIT',
                   'CRÉDIT', 'CREDIT', 'DATE', 'VALEUR', 'LIBEL', 'EXTRAIT'):
            if lbl_up.startswith(kw):
                return True
        # "Page N sur N" — mot isolé SUR/PAGE uniquement s'il est le seul mot significatif
        if re.match(r'^PAGE\s+\d+\s+SUR\s+\d+$', lbl_up):
            return True
        return False

    # ── Extraction solde final depuis le texte ────────────────────────────────
    # "Solde (XOF) au 31/01/2024 : 8 728 070"
    full_text = ' '.join(pages_text)
    m_bal = re.search(
        r'Solde\s+\([A-Z]+\)\s+au\s+\d{2}/\d{2}/\d{4}\s*:\s*([\d\s]+)',
        full_text, re.IGNORECASE
    )
    if m_bal:
        raw_bal = re.sub(r'\s+', '', m_bal.group(1))
        try:
            info['balance_close'] = float(raw_bal)
        except ValueError:
            pass

    # ── Extraction RIB/IBAN BSIC ─────────────────────────────────────────────
    # Format numéro de compte : "01001-00100029193-76 XOF"
    #   → Guichet=01001, Compte=00100029193, Clé=76
    # IBAN collé possible : "Code Iban : SN08SN11101001000100029193"

    m_compte = re.search(
        r'Num[eé]ro\s+de\s+compte\s*:\s*(\d{5})-(\d{8,14})-(\d{2})',
        full_text, re.IGNORECASE
    )
    if m_compte:
        info['_rib_agency']  = m_compte.group(1)   # ex: 01001
        info['_rib_account'] = m_compte.group(2)   # ex: 00100029193
        info['_rib_key']     = m_compte.group(3)   # ex: 76
        if not info.get('_rib_bank'):
            iban_raw = re.sub(r'\s+', '', info.get('iban', '')).upper()
            if re.match(r'^[A-Z]{2}\d{2}', iban_raw) and len(iban_raw) >= 9:
                info['_rib_bank'] = iban_raw[4:9]
            else:
                m_ib = re.search(r'Code\s+[Ii]ban\s*:\s*([A-Z]{2}\d{2}[A-Z0-9]{3,5})',
                                  full_text, re.IGNORECASE)
                if m_ib:
                    info['_rib_bank'] = m_ib.group(1)[4:9]

    # Valider/compléter l'IBAN depuis "Code Iban : SN08SN111..."
    if not info.get('iban') or not re.match(r'^[A-Z]{2}\d{2}', re.sub(r'\s+','',info.get('iban','')).upper()):
        m_ci = re.search(r'Code\s+[Ii]ban\s*:\s*([A-Z]{2}\d{2}[A-Z0-9\s]{14,32})',
                          full_text, re.IGNORECASE)
        if m_ci:
            raw = re.sub(r'\s+', '', m_ci.group(1)).upper()
            raw = re.sub(r'[^A-Z0-9]', '', raw)
            if len(raw) >= 15:
                info['iban'] = raw[:28]

    for pw in pages_words:
        rows = group_words_by_row(pw, tol=2.0)  # réduit pour BSIC (lignes serrées)
        i = 0
        while i < len(rows):
            row = rows[i]
            # Date opération : dd/mm/yyyy à x0 ≈ 32 (tolérance 70 pts)
            date_w = [w for w in row if w['x0'] < 70
                      and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
            if not date_w:
                i += 1; continue

            date_str = date_w[0]['text']
            date_ofx = date_full_to_ofx(date_str)

            # Libellé : x0 entre 120 et 335 sur la ligne courante
            label_words = [w for w in row if 120 <= w['x0'] < 335]
            label = ' '.join(w['text'] for w in label_words).strip()

            # Si le libellé est vide, chercher dans la ou les lignes précédentes
            # (cas BSIC : "Virmt fav. ALIOS FINANCE" sur la ligne juste avant la date)
            if not label and i > 0:
                # Chercher jusqu'à 3 lignes en arrière (sans date op)
                for k in range(i - 1, max(i - 4, -1), -1):
                    prev_row = rows[k]
                    prev_date = [w for w in prev_row if w['x0'] < 70
                                 and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
                    if prev_date:
                        break  # une autre date op → stop
                    prev_label_words = [w for w in prev_row if 120 <= w['x0'] < 335]
                    prev_label = ' '.join(w['text'] for w in prev_label_words).strip()
                    if prev_label and not _is_skip_label(prev_label):
                        label = prev_label
                        break

            # Récupérer les lignes de continuation qui suivent (sans date op)
            j = i + 1
            memo_parts = []
            while j < len(rows):
                next_row = rows[j]
                next_date = [w for w in next_row if w['x0'] < 70
                             and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
                if next_date:
                    break
                cont_words = [w for w in next_row if 120 <= w['x0'] < 335]
                cont = ' '.join(w['text'] for w in cont_words).strip()
                if cont and not _is_skip_label(cont):
                    memo_parts.append(cont)
                j += 1
            i = j

            if not label or _is_skip_label(label):
                continue
            if re.match(r'^[\d\s/\-]+$', label):
                continue

            # Débit (x0 ≈ 355-415) / Crédit (x0 ≈ 440-510)
            debit_words  = [w for w in row if 340 <= w['x0'] < 430]
            credit_words = [w for w in row if 430 <= w['x0'] < 515]

            debit_amt  = _uba_join_amount(debit_words)
            credit_amt = _uba_join_amount(credit_words)

            name, memo = smart_label(label, memo_parts)
            if debit_amt and debit_amt > 0:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt and credit_amt > 0:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))

    # ── Fallback texte brut (PDF scanné via OCR ou pages_words vides) ──────────
    # Quand pages_words est vide (scan OCR), on parse le texte ligne par ligne.
    # Format BSIC OCR attendu par ligne :
    #   "01/10/2025 30/09/2025 RET ESP CP 2149951 MESSIN DRA   80 000   9 262 621"
    #   "07/10/2025 07/10/2025 V/V FAC FV2025 09 000260        867 300   948 931"
    # → date_op  date_val  libellé  [débit ou crédit]  solde
    if not txns:
        full_text_ocr = '\n'.join(pages_text)
        # Pré-traitement : normaliser espaces insécables
        full_text_ocr = full_text_ocr.replace('\xa0', ' ').replace('\u202f', ' ')

        # Extraire toutes les lignes de transaction + la suivante (solde précédent
        # pour déduire le sens débit/crédit)
        lines = full_text_ocr.split('\n')
        prev_solde = None

        for idx, line in enumerate(lines):
            line = line.strip()
            # Doit commencer par date_op dd/mm/yyyy
            m_date = re.match(r'^(\d{2}/\d{2}/\d{4})', line)
            if not m_date:
                continue
            date_str = m_date.group(1)
            date_ofx = date_full_to_ofx(date_str)

            line_up = line.upper()
            if any(kw in line_up for kw in ('SOLDE', 'TOTAL', 'LIBELLÉ', 'LIBELLE',
                                             'DÉBIT', 'DEBIT', 'CRÉDIT', 'CREDIT',
                                             'EXTRAIT', 'PAGE', 'PÉRIODE', 'CODE CLIENT')):
                continue

            # Extraire tous les montants de la ligne (≥ 100 pour ignorer numéros)
            # On cherche des nombres avec éventuels espaces comme séparateurs de milliers
            raw_amounts = re.findall(
                r'\b(\d{1,3}(?:\s\d{3})*(?:,\d{2})?)\b', line)
            amounts_vals = []
            for a in raw_amounts:
                v = parse_amount(a.replace(' ', '.'))
                if v is not None and v >= 100:
                    amounts_vals.append(v)

            if not amounts_vals:
                continue

            # Libellé : retirer date_op + date_val, puis enlever les montants de fin
            label_part = re.sub(r'^\d{2}/\d{2}/\d{4}\s*', '', line)    # date op
            label_part = re.sub(r'^\d{2}/\d{2}/\d{4}\s*', '', label_part)  # date val
            # Retirer les montants numériques de fin de chaîne
            label_part = re.sub(r'[\d\s,]+$', '', label_part).strip()
            label_part = clean_label(label_part)

            if not label_part or len(label_part) < 3 or _is_skip_label(label_part):
                continue

            # ── Déduire débit/crédit ─────────────────────────────────────────
            # Méthode 1 : variation de solde
            # Format : libellé  [montant_op]  solde_courant
            # Si on a ≥ 2 montants : le dernier est le solde, l'avant-dernier est l'opération
            # Si solde_courant < solde_précédent → débit, sinon crédit
            is_credit = None
            if len(amounts_vals) >= 2:
                solde_courant = amounts_vals[-1]
                montant_op    = amounts_vals[-2]
                if prev_solde is not None:
                    diff = solde_courant - prev_solde
                    # Tolérance de 10 XOF pour arrondi OCR
                    if diff > 10:
                        is_credit = True
                    elif diff < -10:
                        is_credit = False
                prev_solde = solde_courant
                amt = montant_op
            else:
                amt = amounts_vals[0]

            # Méthode 2 : mots-clés du libellé (si méthode 1 insuffisante)
            if is_credit is None:
                credit_kw = ('V/V', 'FAC FV', 'VIREMENT ENTRANT', 'REMISE',
                             'VERSEMENT', 'AVOIR', 'RETOUR', 'REMBOURSEMENT',
                             'VIREMENT REÇU', 'CREDIT')
                debit_kw  = ('VIREMENT W', 'FACTURE ARC', 'CHQ', 'PREST',
                             'RET ESP', 'FRAIS', 'RETRAIT', 'COMMISSION',
                             'COTISATION', 'ABONNEMENT', 'TENUE DE COMPTE',
                             'SOUSCRIPTION', 'SMS')
                lbl_up2 = label_part.upper()
                if any(kw in lbl_up2 for kw in credit_kw):
                    is_credit = True
                elif any(kw in lbl_up2 for kw in debit_kw):
                    is_credit = False
                else:
                    is_credit = False  # défaut : débit

            signed_amt = amt if is_credit else -amt
            txns.append(_make_txn(date_ofx, signed_amt, label_part))

    # Fallback universel si toujours rien trouvé
    if not txns and _pdf_path and Path(_pdf_path).exists():
        return _universal_parse_path(_pdf_path, pages_text)
    return info, [t for t in txns if t is not None]


# ── BIS (Banque Islamique du Sénégal) : format iMAL*CSM
# Positions mesurées sur PDF réel :
#   Date+Valeur : x0 ≈ 19-103 (format "dd/mm/yyyydd/mm/yyyy" concaténé ou simple)
#   No Trs      : x0 ≈ 111-152
#   Description : x0 ≈ 157-400
#   Mnt Crédit  : x0 ≈ 320-415  ('0' = colonne vide)
#   Mnt Débit   : x0 ≈ 415-500  ('0' = colonne vide)
#   Solde       : x0 ≈ 509+
def parse_bis(pages_words, pages_text, _pdf_path=''):
    info = _afr_header(pages_text)
    txns = []
    SKIP = {'TOTAL','SOLDE','DATE','VALEUR','NO TRS','DESCRIPTION','MNT','DÉBIT',
            'RÉSUMÉ','BANQUE','ISLAMIQUE','ECOLE','CIF','NO. CPTE','DISPONIBLE',
            'Solde','Bénéficiaire','Page','de','**'}

    for pw in pages_words:
        rows = group_words_by_row(pw, tol=5.0)
        for row in rows:
            # Date : token en x0 < 110 commençant par dd/mm/yyyy
            # Peut être concaténé "02/05/202402/05/2024" (date trs + date valeur)
            date_w = None
            for w in row:
                if w['x0'] < 110:
                    m = re.match(r'^(\d{2}/\d{2}/\d{4})', w['text'])
                    if m:
                        date_w = w; break
            if not date_w:
                continue
            date_str = date_w['text'][:10]  # prendre dd/mm/yyyy

            # Ignorer les lignes de pied de page (contenant heure HH:MM:SS)
            if any(re.match(r'^\d{2}:\d{2}:\d{2}$', w['text']) for w in row):
                continue

            # Description : x0 ~157-400
            desc_words = [w for w in row if 150 <= w['x0'] < 400]
            label = ' '.join(w['text'] for w in desc_words).strip()
            if not label:
                continue
            if re.match(r'^[0-9#,\s]+$', label):
                continue
            if any(s in label for s in SKIP):
                continue

            # Crédit (Mnt Cr) : x0 ≈ 320-415  |  Débit (Mnt débit) : x0 ≈ 415-500
            credit_words = [w for w in row if 320 <= w['x0'] < 415]
            debit_words  = [w for w in row if 415 <= w['x0'] < 500]

            credit_raw = ' '.join(w['text'] for w in credit_words).strip()
            debit_raw  = ' '.join(w['text'] for w in debit_words).strip()

            credit_amt = _uba_join_amount(credit_words) if credit_raw and credit_raw != '0' else None
            debit_amt  = _uba_join_amount(debit_words)  if debit_raw  and debit_raw  != '0' else None

            date_ofx = date_full_to_ofx(date_str)
            name, memo = smart_label(label, [])
            if credit_amt and credit_amt > 0:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))
            elif debit_amt and debit_amt > 0:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))

    if not txns and _pdf_path and Path(_pdf_path).exists():
        return _universal_parse_path(_pdf_path, pages_text)
    return info, [t for t in txns if t is not None]


# ── BNDE : format tableau Date | Libellé | Valeur | Débit | Crédit | Solde
# PDF scanned/image → on utilise le texte extrait ligne par ligne
def parse_bnde(pages_words, pages_text, _pdf_path=''):
    info = _afr_header(pages_text)
    # Essai d'abord avec pdfplumber table
    if _pdf_path and Path(_pdf_path).exists():
        result_info, txns = _universal_parse_path(_pdf_path, pages_text)
        if txns:
            result_info.update({k: v for k, v in info.items() if v})
            return result_info, txns

    # Fallback : parsing texte brut
    year = _year_from_text(' '.join(pages_text[:2]))
    txns = []
    SKIP = {'TOTAL','SOLDE','DATE','LIBELLÉ','LIBELLE','VALEUR','DÉBIT','DEBIT',
            'CRÉDIT','CREDIT','A REPORTER','SOLDE À REPORTER','TITULAIRE','RELEVE',
            'VEUILLEZ','PAGE','BNDE','AGENCE','COMPTE','DEVISE','DOMICILIATION',
            'SIÈGE','SOCIAL','RCCM','NINEA'}

    full_text = '\n'.join(pages_text)
    lines = full_text.splitlines()

    # Pattern : ligne avec date DD/MM en colonne gauche
    date_re = re.compile(r'^(\d{2}/\d{2})\s+(.+)')
    amount_re = re.compile(r'([\d\s]+[\.,]\d{3}(?:[\.,]\d{3})*|[\d\s]+)')

    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = date_re.match(line)
        if not m:
            continue
        day_month = m.group(1)  # DD/MM
        rest = m.group(2).strip()

        # Ignorer les lignes d'en-tête ou de solde
        rest_up = rest.upper()
        if any(s in rest_up for s in SKIP):
            continue

        # Chercher les montants dans le reste de la ligne
        # Format typique : LIBELLÉ   [VALEUR]   DÉBIT   CRÉDIT   SOLDE
        amounts = re.findall(r'[\d\s]{1,15}[\.,]\d{3}(?:[\.,]\d{3})*|[\d]+[\.,]\d{2}', rest)
        amounts_parsed = []
        for a in amounts:
            v = parse_amount(a.strip())
            if v is not None and v > 0:
                amounts_parsed.append(v)

        if not amounts_parsed:
            continue

        # Le libellé est le texte avant les chiffres
        label_match = re.match(r'^([A-Za-zÀ-ÿ\s\-\'\./,&:°N°]+)', rest)
        label = label_match.group(1).strip() if label_match else rest[:50].strip()
        if not label or len(label) < 3:
            continue
        if any(s in label.upper() for s in SKIP):
            continue

        date_ofx = f"{year}{day_month[3:5]}{day_month[0:2]}"
        name, memo = smart_label(label, [])

        # Heuristique: si 2 montants ou plus, le premier non-nul = débit ou crédit
        # On n'a pas la position x donc on utilise le contexte
        # Pour BNDE on essaie la colonne valeur (1er) puis débit (2e) puis crédit (3e)
        if len(amounts_parsed) >= 2:
            # Si le libellé contient des mots associés à un crédit
            credit_kw = {'VERSEMENT','VIREMENT','RECU','REMISE','CREDIT','DÉBLOCAGE',
                         'DEBLOCAGE','ANNUL CHQ','RECOUVREMENT','SWIFT'}
            debit_kw  = {'RETRAIT','CHEQ','CHQ','AGIOS','FRAIS','COMMISSION','APPEL',
                         'ABONNEMENT','ROUTAGE','VIREMENT AUTRE','RETRAIT ESP'}
            is_credit = any(k in label.upper() for k in credit_kw)
            is_debit  = any(k in label.upper() for k in debit_kw)
            amt = amounts_parsed[-1]  # dernier montant souvent débit/crédit
            if is_credit and not is_debit:
                txns.append(_make_txn(date_ofx, amt, name, memo))
            elif is_debit and not is_credit:
                txns.append(_make_txn(date_ofx, -amt, name, memo))
            # else: ambigu, on skip
        elif len(amounts_parsed) == 1:
            # Seul montant : impossible de déterminer sens sans position x
            pass

    if not txns and _pdf_path and Path(_pdf_path).exists():
        return _universal_parse_path(_pdf_path, pages_text)
    return info, [t for t in txns if t is not None]


# ── UBA : format Extrait de compte
# Structure réelle : le libellé est sur les lignes PRÉCÉDANT la ligne de date.
# Ligne de date: date (x0<90), date valeur (x0≈269), débit (x0≈330-430), crédit (x0≈430-520), solde (x0≈520+)
def parse_uba(pages_words, pages_text, _pdf_path=''):
    info = _afr_header(pages_text)
    txns = []
    CREDIT_KW = {'MAINLEVEE','BLOCAGE','RECU','SWIFT','DECAISSEMENT','VERSEMENT','REMISE','CREDIT'}
    DEBIT_KW  = {'VISA','FRAIS','COMMISSION','CHEQUE','TOB','ECHEANCE','PRET',
                 'FACTURATION','REDEVANCE','RETRAIT','AGIOS'}
    HEADER_SKIP = {'SOLDE','DÉBIT','DEBIT','CRÉDIT','CREDIT','VALEUR','DATE',
                   'OPÉRATION','OPERATION','INSTR','AGENCE','COMPTE','PÉRIODE',
                   'POUR','RELEVÉ','EXTRAIT','TITULAIRE','RIB','BIC'}

    def _group_num_tokens(words):
        blocks, cur, prev_x1 = [], [], None
        for w in words:
            is_num = bool(re.match(r'^[\d\s,\.]+$', w['text']) and re.search(r'\d', w['text']))
            if not is_num:
                if cur: blocks.append(cur); cur = []
                prev_x1 = None; continue
            if prev_x1 is not None and (w['x0'] - prev_x1) > 35:
                if cur: blocks.append(cur)
                cur = [w]
            else:
                cur.append(w)
            prev_x1 = w.get('x1', w['x0'] + 20)
        if cur: blocks.append(cur)
        return blocks

    for pw in pages_words:
        rows = group_words_by_row(pw, tol=5.0)

        # Identifier les indices des lignes de date (dd/mm/yyyy à x0 < 90)
        date_row_indices = []
        for idx, row in enumerate(rows):
            dw = [w for w in row if w['x0'] < 90
                  and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
            if dw:
                date_row_indices.append(idx)

        for i, date_idx in enumerate(date_row_indices):
            row = rows[date_idx]
            date_str = [w for w in row if w['x0'] < 90
                        and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])][0]['text']

            # Libellé = lignes entre date row précédente et celle-ci
            prev_date_idx = date_row_indices[i - 1] if i > 0 else -1
            label_parts = []
            for j in range(prev_date_idx + 1, date_idx):
                r = rows[j]
                lw = [w for w in r if 115 <= w['x0'] < 410]
                text = ' '.join(w['text'] for w in lw).strip()
                if not text or re.match(r'^[\d/,\.\s]+$', text):
                    continue
                text_up = text.upper()
                if any(s in text_up for s in HEADER_SKIP):
                    continue
                if re.match(r'^[au\s\d/]+$', text, re.IGNORECASE):
                    continue
                label_parts.append(text)
            label = ' '.join(label_parts).strip()

            if not label:
                row_label = ' '.join(w['text'] for w in row if 80 <= w['x0'] < 260).strip()
                label = row_label

            if not label:
                continue
            label_up = label.upper()
            has_useful = bool(re.search(r'[A-Za-zÀ-ÿ]{3,}', label))
            if not has_useful:
                continue

            # Montants sur la ligne de date (à droite de x0≥250)
            right_words = sorted([w for w in row if w['x0'] >= 250], key=lambda w: w['x0'])
            blocks = _group_num_tokens(right_words)
            numeric_blocks = []
            for b in blocks:
                joined = ''.join(w['text'] for w in b)
                if re.match(r'^\d{2}/\d{2}/\d{4}$', joined):
                    continue
                v = _uba_join_amount(b)
                if v is not None:
                    numeric_blocks.append((v, b[0]['x0']))

            if not numeric_blocks:
                continue

            # Débit x0≈330-430, Crédit x0≈430-520, Solde x0≈520+
            debit_cands  = [(v, x) for v, x in numeric_blocks if 330 <= x < 430]
            credit_cands = [(v, x) for v, x in numeric_blocks if 430 <= x < 520]

            debit_amt  = debit_cands[0][0]  if debit_cands  else None
            credit_amt = credit_cands[0][0] if credit_cands else None

            if debit_amt is None and credit_amt is None and numeric_blocks:
                val, x0 = numeric_blocks[0]
                is_credit = any(k in label_up for k in CREDIT_KW)
                is_debit  = any(k in label_up for k in DEBIT_KW)
                if is_credit and not is_debit:
                    credit_amt = val
                elif is_debit and not is_credit:
                    debit_amt = val

            date_ofx = date_full_to_ofx(date_str)
            name, memo = smart_label(label, [])
            if debit_amt and debit_amt > 0:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt and credit_amt > 0:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))

    if not txns and _pdf_path and Path(_pdf_path).exists():
        return _universal_parse_path(_pdf_path, pages_text)
    return info, [t for t in txns if t is not None]


# ── SG Afrique (SGBS Sénégal) : Date | Libellé | Débit | Crédit
# Particularités : montants XOF entiers sans décimales ('15 400', '369 630'),
# fragmentés en plusieurs tokens par pdfplumber.
def parse_sg_afrique(pages_words, pages_text, _pdf_path=''):
    info = _afr_header(pages_text)
    txns = []
    SKIP = {'TOTAUX DES','NOUVEAU SOLDE','SOLDE PRECEDENT','PROGRAMME DE',
            'RAPPEL DES','MONTANT CUMULE','SOLDE AU','DATE D','DATE DE','LIBELLÉ',
            'TOTAL DES','DÉBIT','CRÉDIT','SOLDE','TOTAL','LIBELLE'}

    # Mots-clés sémantiques pour déduire le sens quand la position seule est ambiguë
    CREDIT_KW = {'VIREMENT','VERSEMENT','REMISE','CREDIT','RECU','SWIFT',
                 'DECAISSEMENT','MAINLEVEE','DEBLOCAGE','TRF-RECU','TRF RECU',
                 'REM.CHQ','REMISE CHQ'}
    DEBIT_KW  = {'REDEVANCE','CHEQUE','COMMISSION','FRAIS','ECHEANCE','PRET',
                 'FACTURATION','RETRAIT','AGIOS','PRELEVEMENT','PRLV',
                 'CHQ COMP','RETRAIT ESPECES','IMPAYE','IBE','ABONNEMENT',
                 'FRAIS TELECOMP'}

    for pw in pages_words:
        rows = group_words_by_row(pw, tol=4.0)
        i = 0
        while i < len(rows):
            row = rows[i]
            # Date : dd/mm/yyyy en x0 < 75
            if not (row and row[0]['x0'] < 75
                    and re.match(r'^\d{2}/\d{2}/\d{4}$', row[0]['text'])):
                i += 1; continue

            label = ' '.join(w['text'] for w in row if 115 <= w['x0'] < 370).strip()
            label_up = label.upper()
            if not label or any(s in label_up for s in SKIP):
                i += 1; continue

            # --- Détection robuste des montants ---
            # SG Sénégal : Date | Date_valeur (79–114) | Libellé (115–370)
            # Débit  ≈ x0 370–420, Crédit ≈ x0 445–490, Solde ≈ x0 520+
            # Zones mesurées sur PDF réel SGBS Sénégal
            right_words = sorted([w for w in row if w['x0'] >= 360],
                                 key=lambda w: w['x0'])

            # Regrouper en blocs contigus (gap > 30px = nouveau bloc)
            def _make_blocks(words, gap=30):
                blocks, cur, prev_x1 = [], [], None
                for w in words:
                    is_num = bool(re.match(r'^[\d,\.]+$', w['text'])
                                  and re.search(r'\d', w['text']))
                    if not is_num:
                        if cur: blocks.append(cur); cur = []
                        prev_x1 = None; continue
                    if prev_x1 is not None and (w['x0'] - prev_x1) > gap:
                        if cur: blocks.append(cur)
                        cur = [w]
                    else:
                        cur.append(w)
                    prev_x1 = w.get('x1', w['x0'] + len(w['text']) * 7)
                if cur: blocks.append(cur)
                return blocks

            blocks = _make_blocks(right_words)

            # Résoudre chaque bloc en valeur numérique
            resolved = []  # liste de (valeur, x0_du_bloc)
            for b in blocks:
                v = _uba_join_amount(b)
                if v is not None:
                    resolved.append((v, b[0]['x0']))

            if not resolved:
                i += 1; continue

            debit_amt = credit_amt = None
            nb = len(resolved)

            # SG Sénégal (SGBS) — positions mesurées sur PDF réel SGBS Oct-2024 :
            #   Débit  : x0 ≈ 430–510  (REDEVANCE, CHEQUE, ECHEANCE, COMMISSION…)
            #   Crédit : x0 ≈ 510–580  (VIREMENT OFMS DECAISSEMENT…)
            #   Solde  : x0 ≈ 580+     (ignoré)
            # La frontière débit/crédit est donc 510 (pas 430).
            # On garde un fallback sémantique pour les cas ambigus (montant unique
            # dont la position ne permet pas de trancher).
            DEBIT_X_MAX  = 510   # colonne débit : [360, 510)
            CREDIT_X_MIN = 510   # colonne crédit : [510, 580)
            SOLDE_X_MIN  = 580   # colonne solde : ≥ 580 (ignorée)

            is_credit_kw = any(k in label_up for k in CREDIT_KW)
            is_debit_kw  = any(k in label_up for k in DEBIT_KW)

            # Filtrer le solde (bloc le plus à droite si x0 ≥ SOLDE_X_MIN)
            # mais seulement quand on a plus de 2 blocs
            candidates = resolved if nb <= 2 else [r for r in resolved if r[1] < SOLDE_X_MIN]
            if not candidates:
                candidates = resolved[:-1] if nb >= 2 else resolved

            if len(candidates) == 1:
                val, x0 = candidates[0]
                if x0 >= CREDIT_X_MIN:
                    credit_amt = val
                elif x0 < DEBIT_X_MAX:
                    # Position clairement dans la colonne débit
                    debit_amt = val
                else:
                    # Zone ambiguë : fallback sémantique
                    if is_credit_kw and not is_debit_kw:
                        credit_amt = val
                    else:
                        debit_amt = val

            elif len(candidates) == 2:
                val_l, x0_l = min(candidates, key=lambda x: x[1])
                val_r, x0_r = max(candidates, key=lambda x: x[1])
                if x0_l < DEBIT_X_MAX and x0_r >= CREDIT_X_MIN:
                    # Colonnes bien séparées : débit gauche, crédit droite
                    debit_amt  = val_l
                    credit_amt = val_r
                elif x0_l >= CREDIT_X_MIN:
                    # Les deux en zone crédit (montant fragmenté)
                    credit_amt = val_l  # le plus à gauche est le 1er token
                else:
                    # Les deux en zone débit
                    debit_amt = val_l

            else:
                # ≥ 3 candidats après filtrage solde : prendre les deux plus distincts
                x_min_c = min(candidates, key=lambda x: x[1])
                x_max_c = max(candidates, key=lambda x: x[1])
                if x_min_c[1] < DEBIT_X_MAX and x_max_c[1] >= CREDIT_X_MIN:
                    debit_amt  = x_min_c[0]
                    credit_amt = x_max_c[0]
                elif x_max_c[1] >= CREDIT_X_MIN:
                    credit_amt = x_max_c[0]
                else:
                    if is_credit_kw and not is_debit_kw:
                        credit_amt = x_max_c[0]
                    else:
                        debit_amt = x_min_c[0]

            # Mémo lignes suivantes
            memo_parts = []
            j = i + 1
            while j < len(rows):
                r2 = rows[j]
                if r2 and r2[0]['x0'] < 75 and re.match(r'^\d{2}/\d{2}/\d{4}$', r2[0]['text']):
                    break
                nl = ' '.join(w['text'] for w in r2 if 115 <= w['x0'] < 370).strip()
                if nl and not any(s in nl.upper() for s in SKIP):
                    memo_parts.append(nl)
                j += 1
            i = j

            date_ofx = date_full_to_ofx(row[0]['text'])
            name, memo = smart_label(label, memo_parts)
            if debit_amt and debit_amt > 0:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt and credit_amt > 0:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))

    if not txns and _pdf_path and Path(_pdf_path).exists():
        return _universal_parse_path(_pdf_path, pages_text)
    return info, [t for t in txns if t is not None]


# ── Parseurs génériques pour les banques africaines moins fréquentes
def _make_african_parser(bank_name):
    def _parser(pages_words, pages_text, _pdf_path=''):
        if _pdf_path and Path(_pdf_path).exists():
            return _universal_parse_path(_pdf_path, pages_text)
        return _afr_header(pages_text), []
    return _parser

parse_cbao      = _make_african_parser('CBAO')
parse_bci       = _make_african_parser('BCI')
parse_orabank   = _make_african_parser('Orabank')
parse_boa       = _make_african_parser('BOA')
parse_atb       = _make_african_parser('ATB')
parse_universal = _make_african_parser('Universal')


# ── CORIS BANK : format « Date | Libellé | Valeur | Débit | Crédit | Solde »
# Montants XOF entiers (sans décimales), dates DD/MM/YYYY.
# Positions mesurées sur relevé Coris Bank Sénégal réel :
#   Date      : x0 < 75
#   Libellé   : x0 ≈ 75–310
#   Valeur    : x0 ≈ 310–390  (date valeur — on l'ignore)
#   Débit     : x0 ≈ 390–470
#   Crédit    : x0 ≈ 470–555
#   Solde     : x0 ≥ 555
def parse_coris(pages_words, pages_text, _pdf_path=''):
    info = _afr_header(pages_text)
    txns = []

    SKIP = {'SOLDE', 'TOTAL', 'TOTAUX', 'DATE', 'LIBELLÉ', 'LIBELLE', 'VALEUR',
            'DÉBIT', 'DEBIT', 'CRÉDIT', 'CREDIT', 'REPORT', 'A REPORTER',
            'NOMBRE', 'MOUVEMENTS', 'ANCIEN', 'NOUVEAU', 'EXTRAIT', 'COMPTE',
            'COMPTES COURANTS', 'RELEVE', 'RELEVÉ', 'PRÉCÉDENT', 'PRECEDENT'}

    CREDIT_KW = {'VERSEMENT', 'VERS.', 'VRT RECU', 'VIREMENT RECU', 'CREDIT',
                 'REMISE', 'SWIFT', 'DÉBLOCAGE', 'DEBLOCAGE', 'RECOUVREMENT',
                 'ANNULATION', 'RETROCESSION'}
    DEBIT_KW  = {'RET ', 'RETRAIT', 'CHEQUE', 'CHEQ', 'FRAIS', 'COMMISSION',
                 'VRT EMIS', 'VIREMENT EMIS', 'PRELEVEMENT', 'AGIOS',
                 'ABONNEMENT', 'IMPAYE', 'PACKAGE', 'CAUTION'}

    for pw in pages_words:
        rows = group_words_by_row(pw, tol=4.0)
        i = 0
        while i < len(rows):
            row = rows[i]

            # Date : DD/MM/YYYY en x0 < 75
            date_w = [w for w in row
                      if w['x0'] < 75 and re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
            if not date_w:
                i += 1; continue

            date_str = date_w[0]['text']

            # Libellé : x0 entre 75 et 310
            label_words = [w for w in row if 75 <= w['x0'] < 310]
            label = ' '.join(w['text'] for w in label_words).strip()
            label_up = label.upper()

            if not label or any(s in label_up for s in SKIP):
                i += 1; continue

            # Collecter les lignes de suite (mémo, libellé multi-lignes)
            memo_parts = []
            j = i + 1
            while j < len(rows):
                r2 = rows[j]
                # Stopper si nouvelle ligne de date
                if r2 and r2[0]['x0'] < 75 and re.match(r'^\d{2}/\d{2}/\d{4}$', r2[0]['text']):
                    break
                nl = ' '.join(w['text'] for w in r2 if 75 <= w['x0'] < 310).strip()
                if nl and not any(s in nl.upper() for s in SKIP):
                    memo_parts.append(nl)
                j += 1
            i = j

            # --- Montants ---
            # Positions mesurées sur PDF réel Coris Bank Sénégal :
            #   Débit  : x0 ≈ 363–420  (ex: '2','100','000' → 2 100 000)
            #   Crédit : x0 ≈ 445–495  (ex: '12','940','000' → 12 940 000)
            #   Solde  : x0 ≈ 520–575  (ex: '27','010','116' → ignoré)
            debit_words  = [w for w in row if 355 <= w['x0'] < 430]
            credit_words = [w for w in row if 440 <= w['x0'] < 510]

            debit_amt  = _uba_join_amount(debit_words)
            credit_amt = _uba_join_amount(credit_words)

            # Fallback sémantique si les colonnes se chevauchent ou sont absentes
            if debit_amt is None and credit_amt is None:
                # Tous les mots numériques à droite
                right_words = sorted([w for w in row if w['x0'] >= 380],
                                     key=lambda w: w['x0'])
                # Exclure les dates (valeur)
                right_words = [w for w in right_words
                               if not re.match(r'^\d{2}/\d{2}/\d{4}$', w['text'])]
                if right_words:
                    # Regrouper en blocs (gap > 25px)
                    blocs, cur, prev_x1 = [], [], None
                    for w in right_words:
                        is_num = bool(re.match(r'^[\d\s,\.]+$', w['text'])
                                      and re.search(r'\d', w['text']))
                        if not is_num:
                            if cur: blocs.append(cur); cur = []
                            prev_x1 = None; continue
                        if prev_x1 is not None and (w['x0'] - prev_x1) > 25:
                            if cur: blocs.append(cur)
                            cur = [w]
                        else:
                            cur.append(w)
                        prev_x1 = w.get('x1', w['x0'] + len(w['text']) * 7)
                    if cur: blocs.append(cur)

                    resolved = [(v, b[0]['x0']) for b in blocs
                                for v in [_uba_join_amount(b)] if v is not None]

                    if resolved:
                        is_credit = any(k in label_up for k in CREDIT_KW)
                        is_debit  = any(k in label_up for k in DEBIT_KW)

                        if len(resolved) == 1:
                            val, x0 = resolved[0]
                            if is_credit and not is_debit:
                                credit_amt = val
                            elif is_debit and not is_credit:
                                debit_amt = val
                            # sinon ambigu sans position fiable
                        elif len(resolved) >= 2:
                            # Exclure le solde (dernier bloc, le plus à droite)
                            candidates = resolved[:-1]
                            if len(candidates) == 1:
                                val, x0 = candidates[0]
                                if x0 >= 440:
                                    credit_amt = val
                                else:
                                    debit_amt = val
                            else:
                                c_left  = min(candidates, key=lambda x: x[1])
                                c_right = max(candidates, key=lambda x: x[1])
                                if is_credit and not is_debit:
                                    credit_amt = c_right[0]
                                elif is_debit and not is_credit:
                                    debit_amt = c_left[0]
                                else:
                                    if c_right[1] >= 440:
                                        credit_amt = c_right[0]
                                    else:
                                        debit_amt = c_left[0]

            date_ofx = date_full_to_ofx(date_str)
            name, memo = smart_label(label, memo_parts)
            if debit_amt and debit_amt > 0:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt and credit_amt > 0:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))

    # Fallback universel si rien trouvé
    if not txns and _pdf_path and Path(_pdf_path).exists():
        return _universal_parse_path(_pdf_path, pages_text)
    return info, [t for t in txns if t is not None]

def parse_ecobank(pages_words, pages_text, _pdf_path=''):
    """
    Parser dédié Ecobank Sénégal.
    Format : Date(DD/MM/YY) | Transaction | DateVal | Débit | Crédit | Solde
    Montants XOF entiers fragmentés. Pas d'IBAN : numéro de compte brut.
    """
    info = _afr_header(pages_text)
    full_text = ' '.join(pages_text)

    # ── Numéro de compte (pas d'IBAN sur ce relevé Ecobank) ─────────────────
    if not info.get('iban') and not info.get('_rib_account'):
        m_cpte = re.search(
            r'Num[eé]ro\s+de\s+compte\s*[:\-]?\s*([\d]{6,20})',
            full_text, re.IGNORECASE
        )
        if m_cpte:
            num = m_cpte.group(1).strip()
            info['iban']         = num
            info['_rib_bank']    = '00000'
            info['_rib_agency']  = '00000'
            info['_rib_account'] = num
            info['_rib_key']     = ''

    # ── Période : "Du 01/01/2025 Au 31/01/2025" ──────────────────────────────
    if not info.get('period_start'):
        m_per = re.search(
            r'(?:Du|du)\s+(\d{2}/\d{2}/\d{4})\s+(?:Au|au)\s+(\d{2}/\d{2}/\d{4})',
            full_text
        )
        if m_per:
            info['period_start'] = m_per.group(1)
            info['period_end']   = m_per.group(2)

    # ── Solde de clôture ─────────────────────────────────────────────────────
    m_bal = re.search(r'Solde\s+de\s+cl[oô]ture\s+([\d\s]+)', full_text, re.IGNORECASE)
    if m_bal:
        raw = re.sub(r'\s+', '', m_bal.group(1))
        try: info['balance_close'] = float(raw)
        except ValueError: pass

    txns = []
    year = _year_from_text(full_text)

    SKIP_UP = {'TOTAL','SOLDE','DATE','DATEVAL','TRANSACTION',
               'DÉBIT','DEBIT','CRÉDIT','CREDIT','PERIODE','PÉRIODE'}

    def _eco_date(row):
        for w in row:
            if w['x0'] < 65 and re.match(r'^\d{2}/\d{2}/(\d{2}|\d{4})$', w['text']):
                return w['text']
        return ''

    def _eco_date_ofx(ds):
        p = ds.split('/')
        if len(p) == 3:
            dd,mm,yy = p
            yr = (2000+int(yy)) if len(yy)==2 and int(yy)<=30 else (1900+int(yy)) if len(yy)==2 else int(yy)
            return f"{yr}{mm.zfill(2)}{dd.zfill(2)}"
        return str(year)+'0101'

    for pw in pages_words:
        rows = group_words_by_row(pw, tol=4.0)
        i = 0
        while i < len(rows):
            row = rows[i]
            date_str = _eco_date(row)
            if not date_str:
                i += 1; continue
            label_words = [w for w in row if 65 <= w['x0'] < 330]
            label = ' '.join(w['text'] for w in label_words).strip()
            label_up = label.upper()
            if not label or any(s in label_up for s in SKIP_UP) or re.match(r'^[\d\s/\-]+$', label):
                i += 1; continue
            memo_parts = []
            j = i + 1
            while j < len(rows):
                r2 = rows[j]
                if _eco_date(r2): break
                cont = ' '.join(w['text'] for w in r2 if 65 <= w['x0'] < 330).strip()
                if cont and not any(s in cont.upper() for s in SKIP_UP):
                    memo_parts.append(cont)
                j += 1
            i = j
            # Colonnes mesurées sur PDF Ecobank Sénégal
            debit_words  = [w for w in row if 390 <= w['x0'] < 465]
            credit_words = [w for w in row if 465 <= w['x0'] < 535]
            debit_amt  = _uba_join_amount(debit_words)
            credit_amt = _uba_join_amount(credit_words)
            # Montant négatif dans colonne débit = remboursement (crédit)
            if debit_amt is None:
                raw_d = ' '.join(w['text'] for w in debit_words)
                if '- ' in raw_d or raw_d.strip().startswith('-'):
                    nums = re.sub(r'[^\d]', '', raw_d)
                    try: credit_amt = float(nums); debit_amt = None
                    except: pass
            date_ofx = _eco_date_ofx(date_str)
            name, memo = smart_label(label, memo_parts)
            if debit_amt and debit_amt > 0:
                txns.append(_make_txn(date_ofx, -debit_amt, name, memo))
            elif credit_amt and credit_amt > 0:
                txns.append(_make_txn(date_ofx, credit_amt, name, memo))

    # Fallback universel si rien extrait
    if not txns and _pdf_path and Path(_pdf_path).exists():
        _, txns2 = _universal_parse_path(_pdf_path, pages_text)
        if txns2:
            return info, txns2

    return info, [t for t in txns if t is not None]


# ════════════════════════════════════════════════════════════════════════════
# DEVISE & LABELS
# ════════════════════════════════════════════════════════════════════════════

BANK_CURRENCY = {
    'QONTO':'EUR','LCL':'EUR','CA':'EUR','CE':'EUR','BP':'EUR','CIC':'EUR',
    'CGD':'EUR','LBP':'EUR','SG':'EUR','BNP':'EUR','MYPOS':'EUR','SHINE':'EUR',
    'CBAO':'XOF','ECOBANK':'XOF','BCI':'XOF','CORIS':'XOF','UBA':'XOF',
    'ORABANK':'XOF','BOA':'XOF','ATB':'TND','SG_AFRIQUE':'XOF','BSIC':'XOF',
    'BIS':'XOF','BNDE':'XOF','UNIVERSAL':'XOF',
}

BANK_LABELS = {
    'QONTO':'Qonto','LCL':'LCL (Crédit Lyonnais)','CA':'Crédit Agricole',
    'CE':"Caisse d'Épargne",'BP':'Banque Populaire','CIC':'CIC',
    'CGD':'Caixa Geral de Depositos','LBP':'La Banque Postale',
    'SG':'Société Générale','BNP':'BNP Paribas','MYPOS':'myPOS',
    'SHINE':'Shine (néo-banque pro)','CBAO':'CBAO (Sénégal)',
    'ECOBANK':'Ecobank','BCI':'BCI','CORIS':'Coris Bank','UBA':'UBA',
    'ORABANK':'Orabank','BOA':'Bank of Africa','ATB':'Arab Tunisian Bank',
    'SG_AFRIQUE':'Société Générale Afrique','BSIC':'BSIC (Sénégal)',
    'BIS':'Banque Islamique du Sénégal','BNDE':'BNDE','UNIVERSAL':'Format universel',
}

AFRICAN_BANKS = {'CBAO','ECOBANK','BCI','CORIS','UBA','ORABANK','BOA','ATB',
                 'SG_AFRIQUE','UNIVERSAL','BSIC','BIS','BNDE'}

PARSERS = {
    'QONTO':parse_qonto,'LCL':parse_lcl,'CA':parse_ca,'CE':parse_ce,
    'BP':parse_bp,'CIC':parse_cic,'CGD':parse_cgd,'LBP':parse_lbp,
    'SG':parse_sg,'BNP':parse_bnp,'MYPOS':parse_mypos,'SHINE':parse_shine,
    'CBAO':parse_cbao,'ECOBANK':parse_ecobank,'BCI':parse_bci,'CORIS':parse_coris,
    'UBA':parse_uba,'ORABANK':parse_orabank,'BOA':parse_boa,'ATB':parse_atb,
    'SG_AFRIQUE':parse_sg_afrique,'UNIVERSAL':parse_universal,
    'BSIC':parse_bsic,'BIS':parse_bis,'BNDE':parse_bnde,
}


# ════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION OFX
# ════════════════════════════════════════════════════════════════════════════

def period_to_ofx(date_str):
    try:
        p = date_str.split('/')
        return f"{p[2]}{p[1].zfill(2)}{p[0].zfill(2)}"
    except:
        return datetime.now().strftime('%Y%m%d')

def _clamp_balance_for_ofx(bal: float) -> tuple:
    """
    Quadra/Cegid limite le champ BALAMT à 11 caractères total
    (signe éventuel + chiffres + point + 2 décimales).
    - Avec décimales : max 9 999 999.99  (format "9999999.99"  = 10 chars)
    - Sans décimales : max 999 999 999   (format "999999999"   = 9 chars)
    En pratique Quadra bloque à partir de 10 chiffres entiers (ex: 1 000 000 000).
    On tronque à 9 999 999.99 pour laisser une marge.
    Retourne (valeur_clampée, True si tronquée).
    """
    QUADRA_MAX = 9_999_999_999.99   # ~10 milliards — limite sûre Quadra
    if abs(bal) > QUADRA_MAX:
        sign = -1 if bal < 0 else 1
        return sign * QUADRA_MAX, True
    return bal, False


def _format_balamt(bal: float, currency: str) -> str:
    """
    Formate le solde pour le champ BALAMT OFX.
    - XOF / devises sans centimes : entier sans décimales
    - EUR et autres : 2 décimales
    Quadra accepte mieux les entiers pour les devises africaines.
    """
    if currency in ('XOF', 'XAF', 'GNF', 'MGA'):
        return f"{int(round(bal))}"
    return f"{bal:.2f}"


def generate_ofx(info, txns, target='quadra', currency='EUR'):
    iban_full = info.get('iban', '') or ''
    bid, brid, aid = iban_to_rib(iban_full, info=info)

    # ── ACCTID pour Quadra / Money ────────────────────────────────────────────
    # Priorite :
    #   1. RIB extrait directement du PDF :
    #      - Banques africaines (IBAN >= 26 chars ou IBAN alphanum) : ACCTID =
    #        numéro de compte seul (ce que Money utilise pour identifier le client)
    #      - Banques FR (IBAN=27 chars, tout numérique) : ACCTID = Banque+Guichet+Compte+Cle
    #   2. IBAN court (<= 22 chars) -> utiliser l'IBAN compact directement.
    #   3. IBAN long (> 22 chars, UEMOA SN=28...) sans _rib_account -> BBAN compte
    #   4. Fallback fragment iban_to_rib().
    iban_is_real  = bool(re.match(r'^[A-Z]{2}\d{2}', iban_full.replace(' ','')))
    iban_compact  = re.sub(r'\s+', '', iban_full)
    # Detecter si c'est une banque africaine : IBAN > 25 chars OU BBAN contient des lettres
    bban_part = iban_compact[4:] if len(iban_compact) > 4 else ''
    is_african_iban = len(iban_compact) > 25 or bool(re.search(r'[A-Z]', bban_part))

    if info.get('_rib_account'):
        bank = info.get('_rib_bank', '') or ''
        if is_african_iban or not iban_is_real:
            # Banque africaine OU num compte brut (pas d'IBAN) :
            # Quadra reconnait le compte via le numero seul
            acctid = info['_rib_account']
        elif bank and bank != '00000':
            # Banque francaise avec RIB complet : Banque+Guichet+Compte+Cle
            acctid = (bank +
                      info.get('_rib_agency','') +
                      info.get('_rib_account','') +
                      info.get('_rib_key',''))
        else:
            acctid = info['_rib_account']
    elif iban_is_real and len(iban_compact) <= 22:
        # IBAN court (FR, BE, NL...) -> IBAN compact
        acctid = iban_compact
    elif iban_is_real and len(iban_compact) > 22:
        # IBAN long (UEMOA SN=28...) -> extraire le numero de compte du BBAN
        # Format BCEAO : CC(2)+KK(2)+Banque(5)+Agence(5)+Compte(11-12)+Cle(2)
        bban = iban_compact[4:]
        compte_bceao = bban[10:-2] if len(bban) >= 13 else bban[10:]
        acctid = compte_bceao if compte_bceao else iban_compact[:22]
    else:
        # Fallback fragment RIB
        acctid = aid

    # ── ACCTID : limiter a 22 caracteres (limite OFX SGML) ───────────────────
    acctid = acctid[:22] if acctid else '0000000000'

    # ── BANKID / BRANCHID : purger les caractères non-alphanumériques ─────────
    bid  = re.sub(r'[^A-Z0-9]', '', bid.upper())[:11]  if bid  else '00000'
    brid = re.sub(r'[^A-Z0-9]', '', brid.upper())[:11] if brid else '00000'

    ds  = period_to_ofx(info.get('period_start',''))
    de  = period_to_ofx(info.get('period_end',''))
    dn  = datetime.now().strftime('%Y%m%d%H')

    # ── Solde : limité à 13 chiffres pour compatibilité Quadra/Cegid ──────────
    bal_raw = info.get('balance_close', 0.0)
    bal, _bal_truncated = _clamp_balance_for_ofx(bal_raw)

    memo_carries_label = target in ('myunisoft','sage','ebp')
    lines = [
        'OFXHEADER:100','DATA:OFXSGML','VERSION:102','SECURITY:NONE',
        'ENCODING:USASCII','CHARSET:1252','COMPRESSION:NONE',
        'OLDFILEUID:NONE','NEWFILEUID:NONE',
        '<OFX>','<SIGNONMSGSRSV1>','<SONRS>','<STATUS>',
        '<CODE>0','<SEVERITY>INFO','</STATUS>',
        f'<DTSERVER>{dn}','<LANGUAGE>FRA',
        '</SONRS>','</SIGNONMSGSRSV1>',
        '<BANKMSGSRSV1>','<STMTTRNRS>','<TRNUID>00',
        '<STATUS>','<CODE>0','<SEVERITY>INFO','</STATUS>',
        '<STMTRS>',f'<CURDEF>{currency}','<BANKACCTFROM>',
        f'<BANKID>{bid}',f'<BRANCHID>{brid}',
        f'<ACCTID>{acctid}','<ACCTTYPE>CHECKING','</BANKACCTFROM>',
        '<BANKTRANLIST>',f'<DTSTART>{ds}',f'<DTEND>{de}',
    ]
    for t in txns:
        name = t['name']
        memo = t.get('memo', '') or ''
        if memo_carries_label:
            name_tag = name
            memo_tag = (name + ' | ' + memo) if memo else name
        else:
            name_tag = name
            memo_tag = memo
        lines += [
            '<STMTTRN>',
            f"<TRNTYPE>{t['type']}",
            f"<DTPOSTED>{t['date']}",
            f"<TRNAMT>{_format_balamt(t['amount'], currency)}",
            f"<FITID>{t['fitid']}",
            '<NAME>' + name_tag,
            '<MEMO>' + memo_tag,
            '</STMTTRN>',
        ]
    bal_fmt = _format_balamt(bal, currency)
    # TRNAMT : aussi entier pour devises sans centimes
    lines += [
        '</BANKTRANLIST>',
        f'<LEDGERBAL>',f'<BALAMT>{bal_fmt}',f'<DTASOF>{dn}','</LEDGERBAL>',
        f'<AVAILBAL>',f'<BALAMT>{bal_fmt}',f'<DTASOF>{dn}','</AVAILBAL>',
        '</STMTRS>','</STMTTRNRS>','</BANKMSGSRSV1>','</OFX>',
    ]
    return '\n'.join(lines) + '\n'


# ════════════════════════════════════════════════════════════════════════════
# FONCTION PRINCIPALE DE CONVERSION (avec cache Streamlit)
# ════════════════════════════════════════════════════════════════════════════

def _pdf_to_images_base64(pdf_path, dpi=200):
    """Convertit chaque page du PDF en image base64 via PyMuPDF."""
    if not _FITZ_OK:
        return []
    images = []
    doc = fitz.open(pdf_path)
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    for page in doc:
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        images.append(base64.b64encode(img_bytes).decode())
    doc.close()
    return images


def _ocr_via_claude(pdf_path):
    """
    OCR de secours : envoie les pages du PDF (images) à l'API Anthropic
    et retourne une liste de textes (une entrée par page).
    Nécessite PyMuPDF (fitz). Utilise l'endpoint /v1/messages sans clé API
    (le proxy Streamlit Cloud l'injecte automatiquement via ANTHROPIC_API_KEY).
    """
    images_b64 = _pdf_to_images_base64(pdf_path)
    if not images_b64:
        return []

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        logger.warning("ANTHROPIC_API_KEY absent — OCR via Claude impossible")
        return []

    pages_text = []
    for img_b64 in images_b64:
        payload = {
            "model": "claude-opus-4-5",
            "max_tokens": 2000,
            "messages": [{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": img_b64
                        }
                    },
                    {
                        "type": "text",
                        "text": (
                            "Ceci est un relevé bancaire scanné. "
                            "Transcris EXACTEMENT tout le texte visible, "
                            "ligne par ligne, en conservant la structure des colonnes "
                            "(date, libellé, débit, crédit, solde). "
                            "Ne reformate pas, ne résume pas. "
                            "Utilise des espaces pour séparer les colonnes. "
                            "Réponds uniquement avec le texte transcrit."
                        )
                    }
                ]
            }]
        }
        try:
            data = json.dumps(payload).encode("utf-8")
            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=data,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                },
                method="POST"
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = json.loads(resp.read().decode())
                page_text = result["content"][0]["text"]
                pages_text.append(page_text)
        except Exception as exc:
            logger.warning("OCR Claude page échouée : %s", exc)
            pages_text.append("")
    return pages_text


@st.cache_data(show_spinner=False)
def process_pdf(file_bytes: bytes, filename: str, force_ocr: bool = False):
    """
    Traite un PDF (bytes) et retourne (bank, info, txns, error).
    Mis en cache par Streamlit — évite de ré-analyser si le fichier n'a pas changé.
    Si force_ocr=True, ignore la détection automatique et active l'OCR directement.
    """
    if not _PDFPLUMBER_OK:
        return None, {}, [], "pdfplumber n'est pas installé. Ajoutez-le à requirements.txt."

    # Écrire dans un fichier temporaire (pdfplumber a besoin d'un chemin)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        pages_words = extract_words_by_page(tmp_path)
        pages_text  = extract_text_by_page(tmp_path)
        _ocr_used = None

        if force_ocr or not _pdf_has_text(pages_text):
            # Essai 1 : Tesseract (si disponible localement)
            if _OCR_AVAILABLE:
                pages_text  = _ocr_pdf(tmp_path)
                pages_words = []
                _ocr_used = 'easyocr' if _EASYOCR_OK else 'tesseract'
            # Essai 2 : OCR via API Claude Vision (PyMuPDF requis)
            elif _FITZ_OK and os.environ.get("ANTHROPIC_API_KEY"):
                logger.info("PDF scanné détecté — OCR via Claude Vision")
                pages_text  = _ocr_via_claude(tmp_path)
                pages_words = []
                _ocr_used = 'claude_vision'
                if not _pdf_has_text(pages_text, min_chars=20):
                    return None, {}, [], (
                        "OCR via Claude Vision n'a pas pu extraire de texte. "
                        "Vérifiez la qualité du scan."
                    )
            else:
                return None, {}, [], (
                    "⚠️ PDF scanné (image uniquement) — aucun texte extractible. "
                    "Solutions : (1) installez Tesseract + pytesseract + pdf2image, "
                    "ou (2) installez PyMuPDF (pip install pymupdf) et définissez "
                    "ANTHROPIC_API_KEY pour activer l'OCR via Claude Vision."
                )

        bank = detect_bank(pages_text)

        if bank in AFRICAN_BANKS:
            info, txns = PARSERS[bank](pages_words, pages_text, _pdf_path=tmp_path)
        else:
            info, txns = PARSERS[bank](pages_words, pages_text)

        # ── Fallback OCR automatique si la lecture standard retourne 0 transaction ──
        if not txns and not force_ocr and not _ocr_used:
            logger.info("Lecture standard : 0 transaction — tentative OCR automatique")
            if _OCR_AVAILABLE:
                pages_text_ocr  = _ocr_pdf(tmp_path)
                pages_words_ocr = []
                bank_ocr = detect_bank(pages_text_ocr)
                if bank_ocr in PARSERS:
                    if bank_ocr in AFRICAN_BANKS:
                        info2, txns2 = PARSERS[bank_ocr](pages_words_ocr, pages_text_ocr, _pdf_path=tmp_path)
                    else:
                        info2, txns2 = PARSERS[bank_ocr](pages_words_ocr, pages_text_ocr)
                    if txns2:
                        info, txns, bank, _ocr_used = info2, txns2, bank_ocr, ('easyocr_auto' if _EASYOCR_OK else 'tesseract_auto')
            elif _FITZ_OK and os.environ.get('ANTHROPIC_API_KEY'):
                pages_text_ocr = _ocr_via_claude(tmp_path)
                bank_ocr = detect_bank(pages_text_ocr)
                if bank_ocr in PARSERS:
                    if bank_ocr in AFRICAN_BANKS:
                        info2, txns2 = PARSERS[bank_ocr]([], pages_text_ocr, _pdf_path=tmp_path)
                    else:
                        info2, txns2 = PARSERS[bank_ocr]([], pages_text_ocr)
                    if txns2:
                        info, txns, bank, _ocr_used = info2, txns2, bank_ocr, 'claude_vision_auto'

        if _ocr_used:
            info['_ocr_mode'] = _ocr_used
        return bank, info, txns, None

    except Exception as e:
        logger.error("Erreur traitement %s : %s", filename, e, exc_info=True)
        return None, {}, [], str(e)
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _generate_excel_export(txns, info, bank_label, currency):
    """
    Génère un fichier Excel professionnel (.xlsx) depuis les transactions.
    Inclut : onglet Transactions formaté, onglet Résumé, styles pro.
    """
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEP1

    wb = openpyxl.Workbook()

    # ── Couleurs ──────────────────────────────────────────────────────────────
    C_HEADER_BG  = '0F2D6B'   # bleu marine
    C_HEADER_FG  = 'FFFFFF'
    C_DEBIT_BG   = 'FFF5F5'   # rouge très pâle
    C_CREDIT_BG  = 'F0FFF4'   # vert très pâle
    C_ALT_BG     = 'F7F9FF'   # bleu pâle alternance
    C_TOTAL_BG   = 'EBF0FF'
    C_TOTAL_FG   = '0F2D6B'
    C_BORDER     = 'CBD5E1'

    thin = Side(style='thin', color=C_BORDER)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Onglet 1 : Transactions ───────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Transactions'

    # En-tête info
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Relevé bancaire — {bank_label}"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=C_HEADER_BG)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    period = f"{info.get('period_start','')} → {info.get('period_end','')}"
    iban_disp = info.get('iban', '')
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Période : {period}   |   IBAN : {iban_disp}   |   Devise : {currency}"
    ws['A2'].font = Font(name='Calibri', size=10, color='5A6B8C')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6  # espace

    # En-têtes colonnes
    headers = ['Date', 'Type', 'Libellé', 'Mémo', f'Débit ({currency})', f'Crédit ({currency})', 'Solde cumulé']
    col_widths = [14, 10, 48, 38, 18, 18, 20]

    hdr_row = 4
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=h)
        cell.font = Font(name='Calibri', size=10, bold=True, color=C_HEADER_FG)
        cell.fill = PatternFill('solid', fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border = border_all
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[hdr_row].height = 22

    # Freeze panes
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)

    # Format montant
    is_xof = currency in ('XOF', 'XAF', 'GNF', 'MGA')
    fmt_num = '#,##0' if is_xof else '#,##0.00'

    cumul = 0.0
    for row_idx, txn in enumerate(txns, hdr_row + 1):
        d = txn['date']
        date_fmt = f"{d[6:8]}/{d[4:6]}/{d[0:4]}"
        is_debit = txn['type'] == 'DEBIT'
        abs_amt = abs(txn['amount'])
        debit_val  = abs_amt if is_debit else None
        credit_val = abs_amt if not is_debit else None
        cumul += txn['amount']

        row_data = [
            date_fmt,
            'Débit' if is_debit else 'Crédit',
            txn['name'],
            txn.get('memo', '') or '',
            debit_val,
            credit_val,
            cumul,
        ]
        # Couleur alternance + débit/crédit
        if is_debit:
            bg = C_DEBIT_BG
        elif row_idx % 2 == 0:
            bg = C_ALT_BG
        else:
            bg = C_CREDIT_BG

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Calibri', size=10)
            cell.fill = PatternFill('solid', fgColor=bg)
            cell.border = border_all
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if col_idx in (5, 6, 7) and val is not None:
                cell.number_format = fmt_num
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_idx].height = 16

    # Ligne TOTAL
    total_row = hdr_row + len(txns) + 1
    total_debit  = sum(abs(t['amount']) for t in txns if t['type'] == 'DEBIT')
    total_credit = sum(t['amount'] for t in txns if t['type'] == 'CREDIT')
    totals_data = ['', 'TOTAL', f"{len(txns)} transaction(s)", '',
                   total_debit, total_credit, cumul]
    for col_idx, val in enumerate(totals_data, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = Font(name='Calibri', size=10, bold=True, color=C_TOTAL_FG)
        cell.fill = PatternFill('solid', fgColor=C_TOTAL_BG)
        cell.border = border_all
        if col_idx in (5, 6, 7) and isinstance(val, float):
            cell.number_format = fmt_num
            cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[total_row].height = 20

    # ── Onglet 2 : Résumé ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title='Résumé')
    summary_rows = [
        ('Banque',         bank_label),
        ('Période',        period),
        ('IBAN',           iban_disp),
        ('Devise',         currency),
        ('Transactions',   len(txns)),
        ('Total Débits',   total_debit),
        ('Total Crédits',  total_credit),
        ('Solde net',      total_credit - total_debit),
        ('Solde d\'ouverture', info.get('balance_open', 0.0)),
        ('Solde de clôture',   info.get('balance_close', 0.0)),
    ]
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 32
    ws2.merge_cells('A1:B1')
    ws2['A1'] = '📊 Résumé du relevé'
    ws2['A1'].font = Font(name='Calibri', size=13, bold=True, color=C_HEADER_BG)
    ws2['A1'].fill = PatternFill('solid', fgColor='EBF0FF')
    ws2['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[1].height = 26

    for r_idx, (label, val) in enumerate(summary_rows, 2):
        la = ws2.cell(row=r_idx, column=1, value=label)
        la.font = Font(name='Calibri', size=10, bold=True, color='374151')
        la.fill = PatternFill('solid', fgColor='F7F9FF' if r_idx % 2 == 0 else 'FFFFFF')
        la.border = border_all
        va = ws2.cell(row=r_idx, column=2, value=val)
        va.font = Font(name='Calibri', size=10)
        va.fill = la.fill
        va.border = border_all
        if isinstance(val, float):
            va.number_format = fmt_num
            va.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ════════════════════════════════════════════════════════════════════════════
# INTERFACE STREAMLIT
# ════════════════════════════════════════════════════════════════════════════

def fmt_amount(amount: float, currency: str) -> str:
    if currency == 'EUR':
        return f"{abs(amount):,.2f} €".replace(",", "\u202f")
    else:
        return f"{abs(amount):,.0f} {currency}".replace(",", "\u202f")


def main():
    st.set_page_config(
        page_title="OFX Bridge — PDF vers OFX",
        page_icon="💱",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── CSS Modern Blue & White ───────────────────────────────────────────────
    st.markdown("""
    <style>
      @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

      /* ── Reset & base ── */
      html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif !important; }
      .stApp { background-color: #f0f4ff !important; }
      .block-container { padding-top: 1.5rem !important; padding-bottom: 3rem !important; max-width: 1100px !important; }

      /* ── Textes globaux ── */
      h1, h2, h3, h4 { color: #0f1f4b !important; font-family: 'Plus Jakarta Sans', sans-serif !important; }
      p, label, .stMarkdown p { color: #4a5568 !important; }

      /* ── Sidebar ── */
      [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0f2d6b 0%, #1a3f8f 100%) !important;
        border-right: none !important;
      }
      [data-testid="stSidebar"] * { color: #e8eeff !important; }
      [data-testid="stSidebar"] h2,
      [data-testid="stSidebar"] h3,
      [data-testid="stSidebar"] strong { color: #ffffff !important; }
      [data-testid="stSidebar"] .stSelectbox label { color: #b8caf5 !important; font-size: 0.82rem !important; font-weight: 600; letter-spacing: 0.06em; text-transform: uppercase; }
      [data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
        background: rgba(255,255,255,0.12) !important;
        border: 1px solid rgba(255,255,255,0.2) !important;
        border-radius: 8px !important;
        color: #fff !important;
      }
      [data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.15) !important; }

      /* ── Metric cards ── */
      [data-testid="metric-container"] {
        background: #ffffff !important;
        border: 1px solid #dce6ff !important;
        border-radius: 14px !important;
        padding: 18px 20px !important;
        box-shadow: 0 2px 12px rgba(15,45,107,0.07) !important;
      }
      [data-testid="stMetricValue"] { color: #0f2d6b !important; font-size: 1.4rem !important; font-weight: 700 !important; }
      [data-testid="stMetricLabel"] { color: #7489b0 !important; font-size: 0.8rem !important; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 0.05em; }

      /* ── Download button ── */
      .stDownloadButton > button {
        background: linear-gradient(135deg, #1a56db 0%, #1e40af 100%) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
        padding: 0.65rem 1.8rem !important;
        letter-spacing: 0.01em !important;
        box-shadow: 0 4px 14px rgba(26,86,219,0.35) !important;
        transition: all 0.2s ease !important;
      }
      .stDownloadButton > button * {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
      }
      .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #1d4ed8 0%, #1e3a8a 100%) !important;
        color: #ffffff !important;
        box-shadow: 0 6px 20px rgba(26,86,219,0.45) !important;
        transform: translateY(-1px) !important;
      }

      /* ── Primary buttons (convert) ── */
      .stButton > button {
        background: linear-gradient(135deg, #1a56db 0%, #1e40af 100%) !important;
        color: #fff !important; border: none !important;
        border-radius: 10px !important; font-weight: 700 !important;
        padding: 0.6rem 1.6rem !important;
        box-shadow: 0 4px 14px rgba(26,86,219,0.3) !important;
      }
      .stButton > button:hover {
        background: linear-gradient(135deg, #1d4ed8 0%, #1e3a8a 100%) !important;
        box-shadow: 0 6px 20px rgba(26,86,219,0.4) !important;
        transform: translateY(-1px) !important;
      }

      /* ── File uploader ── */
      [data-testid="stFileUploader"] {
        background: #ffffff !important;
        border: 2px dashed #93b4f5 !important;
        border-radius: 16px !important;
        padding: 1.5rem !important;
        box-shadow: 0 2px 12px rgba(15,45,107,0.06) !important;
        transition: border-color 0.2s;
      }
      [data-testid="stFileUploader"]:hover { border-color: #1a56db !important; }
      [data-testid="stFileUploader"] label { color: #0f2d6b !important; font-weight: 600 !important; }
      [data-testid="stFileDropzone"] { background: #f5f8ff !important; }

      /* ── Dataframe ── */
      [data-testid="stDataFrame"] {
        border: 1px solid #dce6ff !important;
        border-radius: 12px !important;
        overflow: hidden !important;
        box-shadow: 0 2px 12px rgba(15,45,107,0.06) !important;
      }

      /* ── Alert boxes ── */
      [data-testid="stAlert"] { border-radius: 12px !important; }

      /* ── Separators ── */
      hr { border-color: #dce6ff !important; margin: 1.8rem 0 !important; }

      /* ── Custom components ── */
      .ofx-hero {
        background: linear-gradient(135deg, #0f2d6b 0%, #1a56db 50%, #2563eb 100%);
        border-radius: 20px;
        padding: 2.5rem 2.8rem;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
      }
      .ofx-hero::before {
        content: '';
        position: absolute; top: -40px; right: -40px;
        width: 220px; height: 220px;
        background: rgba(255,255,255,0.06);
        border-radius: 50%;
      }
      .ofx-hero::after {
        content: '';
        position: absolute; bottom: -60px; right: 80px;
        width: 140px; height: 140px;
        background: rgba(255,255,255,0.04);
        border-radius: 50%;
      }
      .ofx-hero h1 {
        color: #ffffff !important;
        font-size: 2rem !important;
        font-weight: 800 !important;
        margin: 0 0 0.4rem 0 !important;
        letter-spacing: -0.02em;
      }
      .ofx-hero p {
        color: rgba(255,255,255,0.78) !important;
        font-size: 1.05rem !important;
        margin: 0 !important;
      }
      .ofx-hero .badges {
        margin-top: 1.2rem;
        display: flex; gap: 10px; flex-wrap: wrap;
      }
      .ofx-hero .badge {
        background: rgba(255,255,255,0.15);
        border: 1px solid rgba(255,255,255,0.25);
        color: #fff !important;
        padding: 4px 14px;
        border-radius: 20px;
        font-size: 0.82rem;
        font-weight: 600;
      }

      .step-card {
        background: #fff;
        border: 1px solid #dce6ff;
        border-radius: 14px;
        padding: 1.2rem 1.5rem;
        display: flex; align-items: flex-start; gap: 14px;
        box-shadow: 0 2px 10px rgba(15,45,107,0.06);
      }
      .step-num {
        background: linear-gradient(135deg, #1a56db, #2563eb);
        color: #fff; font-weight: 800; font-size: 1rem;
        width: 32px; height: 32px; border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        flex-shrink: 0;
      }
      .step-text strong { color: #0f2d6b !important; font-size: 0.95rem; }
      .step-text small { color: #7489b0 !important; font-size: 0.82rem; }

      .bank-badge {
        background: linear-gradient(135deg, #eff6ff, #dbeafe);
        color: #1a56db !important;
        border: 1px solid #bfdbfe;
        padding: 5px 14px; border-radius: 20px;
        font-size: 0.85rem; font-weight: 700;
        display: inline-block; margin-bottom: 10px;
        letter-spacing: 0.01em;
      }

      .result-card {
        background: #ffffff;
        border: 1px solid #dce6ff;
        border-radius: 16px;
        padding: 1.8rem 2rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 4px 20px rgba(15,45,107,0.08);
      }
      .result-card .file-title {
        font-size: 1rem; font-weight: 700; color: #0f2d6b !important;
        display: flex; align-items: center; gap: 8px; margin-bottom: 1rem;
      }

      .info-row {
        background: #f5f8ff;
        border: 1px solid #dce6ff;
        border-radius: 10px;
        padding: 10px 16px;
        margin-bottom: 1rem;
        display: flex; gap: 2rem; flex-wrap: wrap;
      }
      .info-item { font-size: 0.85rem; color: #4a5568 !important; }
      .info-item strong { color: #0f2d6b !important; }

      .security-box {
        background: linear-gradient(135deg, #eff6ff, #f0fdf4);
        border: 1px solid #bfdbfe;
        border-radius: 12px;
        padding: 1rem 1.2rem;
        margin-top: 0.5rem;
      }
      .security-box p { color: #1e40af !important; font-size: 0.88rem !important; margin: 0 !important; }

      .sidebar-label {
        color: rgba(184,202,245,0.9) !important;
        font-size: 0.72rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.09em !important;
        text-transform: uppercase !important;
        margin-bottom: 0.3rem !important;
      }

      .footer-bar {
        background: #fff;
        border: 1px solid #dce6ff;
        border-radius: 12px;
        padding: 1rem 1.5rem;
        display: flex; justify-content: space-between; align-items: center;
        flex-wrap: wrap; gap: 0.5rem;
      }
      .footer-bar span { color: #7489b0 !important; font-size: 0.82rem !important; }
      .footer-bar .highlight { color: #1a56db !important; font-weight: 700 !important; }

      .empty-state {
        background: #fff;
        border: 2px dashed #c3d4f8;
        border-radius: 20px;
        padding: 3rem 2rem;
        text-align: center;
      }
      .empty-state .icon { font-size: 3rem; margin-bottom: 1rem; }
      .empty-state h3 { color: #0f2d6b !important; font-size: 1.2rem !important; margin-bottom: 0.5rem !important; }
      .empty-state p { color: #7489b0 !important; font-size: 0.9rem !important; }

      /* ── Data editor (tableau éditable) ── */
      [data-testid="stDataEditor"] {
        border: 1px solid #dce6ff !important;
        border-radius: 12px !important;
        overflow: hidden !important;
        box-shadow: 0 2px 12px rgba(15,45,107,0.06) !important;
      }
      [data-testid="stDataEditor"] [data-testid="glideDataEditor"] {
        border-radius: 12px !important;
      }
      /* Highlight modifié */
      .edit-notice {
        background: #fffbeb; border: 1px solid #fcd34d;
        border-radius: 10px; padding: 10px 16px;
        font-size: 0.85rem; color: #92400e;
        margin: 0.5rem 0;
      }

      /* Spinner override */
      .stSpinner > div { border-top-color: #1a56db !important; }

      /* Section headers */
      .section-header {
        font-size: 1rem; font-weight: 700; color: #0f2d6b !important;
        margin: 0 0 0.8rem 0; padding-bottom: 0.5rem;
        border-bottom: 2px solid #e8f0ff;
      }
    </style>
    """, unsafe_allow_html=True)

    # ── SIDEBAR ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("""
        <div style="padding: 0.5rem 0 1rem 0;">
          <div style="font-size:1.5rem; font-weight:800; color:#fff; letter-spacing:-0.02em;">💱 OFX Bridge</div>
          <div style="font-size:0.8rem; color:rgba(255,255,255,0.55); margin-top:2px;">v2.1 — Convertisseur PDF → OFX</div>
        </div>
        """, unsafe_allow_html=True)

        st.divider()

        st.markdown('<div class="sidebar-label">⚙️ Logiciel comptable cible</div>', unsafe_allow_html=True)
        target = st.selectbox(
            "Logiciel cible",
            options=["Quadra / Cegid", "MyUnisoft", "Sage", "EBP"],
            index=0,
            label_visibility="collapsed",
            help="Affecte la disposition NAME/MEMO dans le fichier OFX."
        )
        target_map = {"Quadra / Cegid": "quadra", "MyUnisoft": "myunisoft",
                      "Sage": "sage", "EBP": "ebp"}
        target_code = target_map[target]

        st.divider()

        st.markdown('<div class="sidebar-label">🔍 Options OCR</div>', unsafe_allow_html=True)

        # ── Statut OCR ────────────────────────────────────────────────────────
        if _EASYOCR_OK:
            st.markdown(
                "<div style='font-size:0.78rem; background:rgba(167,139,250,0.2); "
                "border:1px solid rgba(167,139,250,0.4); border-radius:8px; padding:6px 10px; "
                "color:#ddd6fe; margin-bottom:8px;'>🔬 <b>EasyOCR</b> actif</div>",
                unsafe_allow_html=True
            )
        elif _TESSERACT_OK:
            st.markdown(
                "<div style='font-size:0.78rem; background:rgba(134,239,172,0.2); "
                "border:1px solid rgba(134,239,172,0.3); border-radius:8px; padding:6px 10px; "
                "color:#bbf7d0; margin-bottom:8px;'>🔍 <b>Tesseract OCR</b> actif</div>",
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                "<div style='font-size:0.78rem; background:rgba(239,68,68,0.2); "
                "border:1px solid rgba(239,68,68,0.3); border-radius:8px; padding:6px 10px; "
                "color:#fca5a5; margin-bottom:8px;'>❌ <b>Aucun OCR</b> — PDF texte uniquement<br>"
                "<span style=\"font-size:0.72rem\">Installez <code>easyocr</code> pour les scans</span></div>",
                unsafe_allow_html=True
            )

        force_ocr = st.checkbox(
            "Forcer l'OCR (relevé scanné)",
            value=False,
            help=(
                "Cochez cette option si votre relevé est un PDF scanné (image) "
                "et que la détection automatique ne trouve aucune transaction. "
                "Nécessite EasyOCR ou Tesseract installé."
            )
        )
        if force_ocr:
            st.markdown(
                "<div style='font-size:0.76rem; color:#fcd34d; margin-top:-0.3rem; margin-bottom:0.4rem;'>"
                "⚡ OCR activé — le traitement sera plus lent.</div>",
                unsafe_allow_html=True
            )

        # ── Mode debug ───────────────────────────────────────────────────────
        mode_debug = st.checkbox(
            "🐛 Mode debug",
            value=False,
            help="Affiche le texte extrait par page, les anomalies de parsing et les données brutes."
        )

        st.divider()

        st.markdown("""
        <div class="security-box" style="background:rgba(255,255,255,0.1);border:1px solid rgba(255,255,255,0.2);">
          <div style="font-size:0.95rem; font-weight:700; color:#fff; margin-bottom:5px;">🔒 100 % local & sécurisé</div>
          <div style="font-size:0.82rem; color:rgba(255,255,255,0.7); line-height:1.5;">
            Vos relevés sont traités <strong style="color:#7dd3fc">directement sur ce serveur</strong>.
            Aucun fichier n'est envoyé vers le cloud. Aucune IA distante. Aucune inscription requise.
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.divider()

        st.markdown("""
        <div style="font-size:0.78rem; color:rgba(255,255,255,0.45); line-height:1.7;">
          <div style="font-weight:700; color:rgba(255,255,255,0.6); margin-bottom:5px; font-size:0.72rem; text-transform:uppercase; letter-spacing:0.08em;">Détection automatique</div>
          Qonto · LCL · Crédit Agricole · Caisse d'Épargne · Banque Populaire · CIC · La Banque Postale · Société Générale · BNP Paribas · CGD · myPOS · Shine · CBAO · Ecobank · BCI · Coris · UBA · Orabank · BOA · ATB · BSIC · BIS · BNDE · + Format universel
        </div>
        """, unsafe_allow_html=True)

    # ── HERO HEADER ───────────────────────────────────────────────────────────
    st.markdown("""
    <div class="ofx-hero">
      <h1>💱 Convertisseur PDF → OFX</h1>
      <p>Importez vos relevés bancaires directement dans Quadra, MyUnisoft, Sage ou EBP — en quelques secondes.</p>
      <div class="badges">
        <span class="badge">✓ Détection automatique de la banque</span>
        <span class="badge">✓ 100 % gratuit &amp; local</span>
        <span class="badge">✓ Multi-fichiers</span>
        <span class="badge">✓ Aucune inscription</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Vérification dépendances
    if not _PDFPLUMBER_OK:
        st.error("❌ **pdfplumber n'est pas installé.** Ajoutez `pdfplumber` à votre `requirements.txt` et relancez l'application.")
        st.stop()

    # ── Comment ça marche ─────────────────────────────────────────────────────
    col_s1, col_s2, col_s3 = st.columns(3)
    with col_s1:
        st.markdown("""<div class="step-card">
          <div class="step-num">1</div>
          <div class="step-text"><strong>Déposez votre PDF</strong><br><small>Relevé bancaire natif ou scanné (OCR)</small></div>
        </div>""", unsafe_allow_html=True)
    with col_s2:
        st.markdown("""<div class="step-card">
          <div class="step-num">2</div>
          <div class="step-text"><strong>Corrigez si nécessaire</strong><br><small>Tableau éditable : date, libellé, montant, type</small></div>
        </div>""", unsafe_allow_html=True)
    with col_s3:
        st.markdown("""<div class="step-card">
          <div class="step-num">3</div>
          <div class="step-text"><strong>Téléchargez l'OFX</strong><br><small>Compatible avec votre logiciel comptable</small></div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

    # ── Upload ────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">📂 Sélection des fichiers PDF</div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "Glissez-déposez un ou plusieurs relevés bancaires au format PDF",
        type=["pdf"],
        accept_multiple_files=True,
        help="PDF natif (texte extractible) recommandé. Les PDF scannés utilisent EasyOCR (précision améliorée sur les chiffres) ou Tesseract comme fallback."
    )

    if not uploaded_files:
        st.markdown("""
        <div class="empty-state">
          <div class="icon">📄</div>
          <h3>Aucun fichier sélectionné</h3>
          <p>Glissez-déposez vos relevés PDF ci-dessus pour démarrer la conversion.<br>
          La banque est détectée automatiquement — aucune configuration requise.</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1: st.metric("Transactions", "—")
        with c2: st.metric("Total Débits", "—")
        with c3: st.metric("Total Crédits", "—")
        st.stop()

    # ── Traitement de chaque fichier ──────────────────────────────────────────
    import pandas as pd

    for uploaded_file in uploaded_files:
        st.markdown(f"<div style='height:0.5rem'></div>", unsafe_allow_html=True)

        file_bytes = uploaded_file.read()

        with st.spinner(f"Analyse de « {uploaded_file.name} »…"):
            bank, info, txns, error = process_pdf(file_bytes, uploaded_file.name, force_ocr=force_ocr)

        # ── Carte de résultat ──────────────────────────────────────────────────
        st.markdown('<div class="result-card">', unsafe_allow_html=True)
        st.markdown(f'<div class="file-title">📄 {uploaded_file.name}</div>', unsafe_allow_html=True)

        if error:
            st.error(f"❌ **Erreur :** {error}")
            st.markdown('</div>', unsafe_allow_html=True)
            continue

        if not txns:
            # Distinguer PDF scanné sans OCR vs. relevé non reconnu
            if not _OCR_AVAILABLE:
                st.warning(
                    "⚠️ **PDF scanné détecté** — Ce relevé est une image (scan). "
                    "Aucun moteur OCR n'est disponible. "
                    "Pour traiter ce fichier, installez **easyocr** (recommandé) ou **pytesseract + tesseract-ocr** "
                    "puis redémarrez l'application."
                )
            else:
                st.warning("⚠️ Aucune transaction détectée. Vérifiez qu'il s'agit bien d'un relevé bancaire.")
            st.markdown('</div>', unsafe_allow_html=True)
            continue

        # Badge banque + infos
        bank_label = BANK_LABELS.get(bank, bank)
        currency   = BANK_CURRENCY.get(bank, 'EUR')
        file_key   = uploaded_file.name.replace('.', '_').replace(' ', '_')

        period_str = ""
        if info.get('period_start') and info.get('period_end'):
            period_str = f"{info['period_start']} → {info['period_end']}"
        iban_detected = info.get('iban', '') or ''

        ocr_badge = ""
        ocr_mode = info.get('_ocr_mode', '')
        if ocr_mode == 'claude_vision' or ocr_mode == 'claude_vision_auto':
            ocr_badge = '<span style="background:#fef3c7;color:#92400e;font-size:0.75rem;font-weight:700;padding:3px 10px;border-radius:20px;border:1px solid #fcd34d">🤖 OCR Claude Vision</span>'
        elif ocr_mode in ('easyocr', 'easyocr_auto'):
            ocr_badge = '<span style="background:#ede9fe;color:#5b21b6;font-size:0.75rem;font-weight:700;padding:3px 10px;border-radius:20px;border:1px solid #c4b5fd">🔬 OCR EasyOCR</span>'
        elif ocr_mode in ('tesseract', 'tesseract_auto'):
            ocr_badge = '<span style="background:#dcfce7;color:#166534;font-size:0.75rem;font-weight:700;padding:3px 10px;border-radius:20px;border:1px solid #86efac">🔍 OCR Tesseract</span>'

        st.markdown(f"""
        <div style="display:flex; align-items:center; gap:12px; margin-bottom:1rem; flex-wrap:wrap;">
          <span class="bank-badge">🏦 {bank_label}</span>
          {ocr_badge}
          <span style="font-size:0.85rem; color:#4a5568">
            {'📅 ' + period_str if period_str else ''}
          </span>
        </div>
        """, unsafe_allow_html=True)

        # ── Champ IBAN/RIB éditable ────────────────────────────────────────────
        iban_key = f"iban_{file_key}"
        if iban_key not in st.session_state:
            st.session_state[iban_key] = iban_detected

        col_iban1, col_iban2 = st.columns([3, 1])
        with col_iban1:
            iban_label = "🔑 RIB (banque guichet compte clé)" if info.get('_rib_bank') else "🔑 IBAN"
            iban_help  = ("RIB extrait du relevé — format : BBBBB GGGGG CCCCCCCCCCC KK\n"
                          "Modifiez si la détection est incorrecte."
                          if info.get('_rib_bank') else
                          "IBAN extrait du relevé. Modifiez si la détection est incorrecte.")
            iban_value = st.text_input(
                iban_label,
                value=st.session_state[iban_key],
                key=f"iban_input_{file_key}",
                help=iban_help,
                placeholder="Ex: 30003 03320 00020641644 69 (RIB) ou FR76 ... (IBAN)",
            )
        with col_iban2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("↺ Réinitialiser", key=f"iban_reset_{file_key}",
                         help="Revenir à la valeur détectée automatiquement"):
                st.session_state[iban_key] = iban_detected
                st.rerun()

        # Mettre à jour info avec l'IBAN/RIB saisi manuellement
        if iban_value != iban_detected:
            iban_clean = iban_value.strip()
            # RIB français : 5+5+11+2 (tout numérique)
            rib_fr  = re.match(r'^(\d{5})\s+(\d{5})\s+(\d{10,12})\s+(\d{2})$', iban_clean)
            # RIB africain : code banque alphanumérique + agence + compte + clé
            # Ex: "SN213 01001 02341624101 33"
            rib_afr = re.match(r'^([A-Z]{0,2}\d{3,5})\s+(\d{4,6})\s+(\d{8,14})\s+(\d{2})$', iban_clean)
            if rib_fr:
                info['_rib_bank']    = rib_fr.group(1)
                info['_rib_agency']  = rib_fr.group(2)
                info['_rib_account'] = rib_fr.group(3)
                info['_rib_key']     = rib_fr.group(4)
                info['iban']         = iban_clean
            elif rib_afr:
                info['_rib_bank']    = rib_afr.group(1)
                info['_rib_agency']  = rib_afr.group(2)
                info['_rib_account'] = rib_afr.group(3)
                info['_rib_key']     = rib_afr.group(4)
                info['iban']         = iban_clean
            else:
                info['iban'] = iban_clean
                # Nettoyer les _rib_* pour forcer iban_to_rib à re-dériver depuis l'IBAN
                info.pop('_rib_bank', None)
                info.pop('_rib_agency', None)
                info.pop('_rib_account', None)
                info.pop('_rib_key', None)
        elif iban_detected:
            info['iban'] = iban_detected

        iban_display = info.get('iban', '') or '—'
        # ── Diagnostic IBAN/RIB ───────────────────────────────────────────────
        rib_bank    = info.get('_rib_bank', '')
        rib_agency  = info.get('_rib_agency', '')
        rib_account = info.get('_rib_account', '')
        rib_key     = info.get('_rib_key', '')

        if rib_bank and rib_account:
            # Afficher décomposition RIB
            st.markdown(
                f"<div style='font-size:0.78rem; color:#7489b0; margin-top:-0.4rem; margin-bottom:0.4rem;'>"
                f"🔑 RIB détecté&nbsp;: "
                f"<code style='color:#1a56db'>Banque&nbsp;<b>{rib_bank}</b></code> "
                f"<code style='color:#1a56db'>Guichet&nbsp;<b>{rib_agency}</b></code> "
                f"<code style='color:#1a56db'>Compte&nbsp;<b>{rib_account}</b></code> "
                f"<code style='color:#1a56db'>Clé&nbsp;<b>{rib_key}</b></code>"
                f"</div>",
                unsafe_allow_html=True
            )
        elif iban_display != '—':
            st.markdown(
                f"<div style='font-size:0.78rem; color:#7489b0; margin-top:-0.4rem; margin-bottom:0.4rem;'>"
                f"🔑 IBAN utilisé dans l'OFX&nbsp;: <code style='color:#1a56db'>{iban_display}</code></div>",
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                "<div style='font-size:0.78rem; color:#e53e3e; margin-top:-0.4rem; margin-bottom:0.4rem;'>"
                "⚠️ <b>Aucun IBAN/RIB détecté</b> — saisissez-le manuellement ci-dessus pour un OFX valide.</div>",
                unsafe_allow_html=True
            )

        # ── Avertissement solde trop grand (Quadra) ───────────────────────────
        bal_raw_check = info.get('balance_close', 0.0)
        _, bal_will_truncate = _clamp_balance_for_ofx(bal_raw_check)
        if bal_will_truncate:
            st.warning(
                f"⚠️ **Solde trop grand pour Quadra/Cegid** — La valeur `{bal_raw_check:,.2f}` "
                f"dépasse la limite du champ BALAMT (~10 milliards). "
                f"La valeur sera plafonnée à `9 999 999 999.99` dans l'OFX. "
                f"Les **transactions individuelles** ne sont pas affectées — l'import reste fonctionnel."
            )

        # ── Métriques ──────────────────────────────────────────────────────────
        total_debit  = sum(abs(t['amount']) for t in txns if t['type'] == 'DEBIT')
        total_credit = sum(t['amount']      for t in txns if t['type'] == 'CREDIT')
        balance      = total_credit - total_debit

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Transactions", f"{len(txns)}")
        with col2:
            st.metric("Total Débits", fmt_amount(total_debit, currency))
        with col3:
            st.metric("Total Crédits", fmt_amount(total_credit, currency))
        with col4:
            delta_label = "positif" if balance >= 0 else "négatif"
            st.metric("Solde net", fmt_amount(abs(balance), currency),
                      delta=f"{'+'  if balance >= 0 else '-'}{fmt_amount(abs(balance), currency)}",
                      delta_color="normal" if balance >= 0 else "inverse")

        # ── Audit qualité ──────────────────────────────────────────────────────
        audit_issues = audit_transactions(txns)
        if audit_issues:
            st.markdown(
                f"<div style='background:#fff7ed; border:1px solid #fed7aa; border-radius:10px; "
                f"padding:10px 16px; margin:0.5rem 0; font-size:0.85rem; color:#9a3412;'>"
                f"⚠️ <strong>{len(audit_issues)} anomalie(s) détectée(s)</strong> dans l'extraction "
                f"— vérifiez le tableau ci-dessous.</div>",
                unsafe_allow_html=True
            )
            if mode_debug:
                with st.expander(f"🔍 Détail des anomalies ({len(audit_issues)})", expanded=True):
                    for issue in audit_issues[:20]:
                        st.markdown(f"- `{issue}`")

        # ── Debug : texte extrait par page ────────────────────────────────────
        if mode_debug:
            # On re-lit le texte depuis le cache (process_pdf retourne txns, pas pages_text)
            # On affiche un résumé de ce qu'on a reçu
            with st.expander("🔍 Debug : texte brut extrait du PDF (via cache)", expanded=False):
                st.markdown(
                    "_Note : le texte affiché ici est celui utilisé par le parseur. "
                    "Pour voir le texte brut page par page, re-uploadez le fichier avec le mode debug activé._"
                )
                st.markdown(f"**Banque détectée :** `{bank}` | **OCR :** `{info.get('_ocr_mode', 'non utilisé')}`")

        # ── Tableau ÉDITABLE ───────────────────────────────────────────────────
        st.markdown(f"""
        <div class="section-header" style="margin-top:1.2rem">
          ✏️ Vérification &amp; correction des transactions
          <span style="font-size:0.78rem; font-weight:500; color:#7489b0; margin-left:10px;">
            Cliquez sur une cellule pour modifier • Supprimer une ligne avec la case à gauche
          </span>
        </div>""", unsafe_allow_html=True)

        # Construction du DataFrame éditable (montants numériques)
        edit_rows = []
        for t in txns:
            d = t['date']
            date_fmt = f"{d[6:8]}/{d[4:6]}/{d[0:4]}"
            is_debit = t['type'] == 'DEBIT'
            edit_rows.append({
                "📅 Date":    date_fmt,
                "Type":       "Débit" if is_debit else "Crédit",
                "Libellé":    t['name'],
                "Mémo":       t.get('memo', '') or '',
                "Montant":    round(abs(t['amount']), 2),
            })

        df_edit = pd.DataFrame(edit_rows)

        edited_df = st.data_editor(
            df_edit,
            use_container_width=True,
            height=min(480, 60 + 38 * len(df_edit)),
            hide_index=False,
            num_rows="dynamic",
            key=f"editor_{file_key}",
            column_config={
                "📅 Date":  st.column_config.TextColumn(
                    "📅 Date", width=105,
                    help="Format JJ/MM/AAAA — modifiable directement",
                ),
                "Type": st.column_config.SelectboxColumn(
                    "Type", width=100,
                    options=["Débit", "Crédit"],
                    required=True,
                    help="Inverser Débit ↔ Crédit si mal détecté",
                ),
                "Libellé": st.column_config.TextColumn(
                    "Libellé", width=260,
                    help="Nom de la transaction exporté dans NAME",
                ),
                "Mémo": st.column_config.TextColumn(
                    "Mémo", width=200,
                    help="Champ MEMO dans l'OFX",
                ),
                "Montant": st.column_config.NumberColumn(
                    f"Montant ({currency})", width=130,
                    min_value=0.0, step=0.01, format="%.2f",
                    help="Valeur absolue — le signe est géré par le Type",
                ),
            },
        )

        # Indicateur de modifications
        orig_hash = str(df_edit.values.tolist())
        edit_hash = str(edited_df.values.tolist())
        has_changes = (orig_hash != edit_hash)
        nb_edited   = len(edited_df)

        if has_changes:
            st.markdown("""
            <div style="background:#fffbeb; border:1px solid #fcd34d; border-radius:10px;
                        padding:10px 16px; margin:0.5rem 0; font-size:0.85rem; color:#92400e;">
              ✏️ <strong>Modifications détectées</strong> — L'OFX sera généré depuis le tableau corrigé ci-dessus.
            </div>""", unsafe_allow_html=True)

        # ── Reconstruction des transactions depuis le tableau édité ────────────
        txns_final = []
        errors_edit = []
        for idx, row in edited_df.iterrows():
            try:
                raw_date  = str(row.get("📅 Date", "")).strip()
                txn_type  = str(row.get("Type", "Débit")).strip()
                label     = str(row.get("Libellé", "")).strip()
                memo      = str(row.get("Mémo", "")).strip()
                montant   = float(row.get("Montant", 0) or 0)

                if not raw_date or not label or montant == 0:
                    continue

                # Conversion date JJ/MM/AAAA → AAAAMMJJ
                date_ofx = date_full_to_ofx(raw_date)
                if not re.match(r'^\d{8}$', date_ofx):
                    errors_edit.append(f"Ligne {idx+1} : date invalide « {raw_date} »")
                    continue

                amount = -montant if txn_type == "Débit" else montant
                txn = _make_txn(date_ofx, amount, label[:64], memo[:128])
                if txn:
                    txns_final.append(txn)
            except Exception as ex:
                errors_edit.append(f"Ligne {idx+1} : {ex}")

        if errors_edit:
            for err in errors_edit:
                st.warning(f"⚠️ {err}")

        # Métriques recalculées sur les données éditées
        if has_changes and txns_final:
            total_debit_e  = sum(abs(t['amount']) for t in txns_final if t['type'] == 'DEBIT')
            total_credit_e = sum(t['amount']      for t in txns_final if t['type'] == 'CREDIT')
            balance_e      = total_credit_e - total_debit_e
            col1e, col2e, col3e, col4e = st.columns(4)
            with col1e: st.metric("Transactions (éditées)", f"{len(txns_final)}")
            with col2e: st.metric("Total Débits",  fmt_amount(total_debit_e,  currency))
            with col3e: st.metric("Total Crédits", fmt_amount(total_credit_e, currency))
            with col4e: st.metric("Solde net", fmt_amount(abs(balance_e), currency),
                                  delta=f"{'+'if balance_e>=0 else '-'}{fmt_amount(abs(balance_e),currency)}",
                                  delta_color="normal" if balance_e >= 0 else "inverse")

        # OFX généré depuis les données (potentiellement éditées)
        txns_for_ofx = txns_final if txns_final else txns

        # ── Téléchargement OFX + Excel ────────────────────────────────────────
        st.markdown('<div class="section-header" style="margin-top:1.2rem">⬇️ Télécharger les fichiers</div>', unsafe_allow_html=True)

        ofx_content = generate_ofx(info, txns_for_ofx, target=target_code, currency=currency)
        ofx_bytes   = ofx_content.encode('latin-1', errors='replace')
        ofx_name    = Path(uploaded_file.name).stem + ".ofx"
        xlsx_name   = Path(uploaded_file.name).stem + ".xlsx"

        # Générer Excel
        try:
            xlsx_bytes = _generate_excel_export(txns_for_ofx, info, bank_label, currency)
            excel_ok = True
        except Exception as exc_xl:
            logger.warning("Export Excel échoué : %s", exc_xl)
            excel_ok = False

        col_dl1, col_dl2, col_dl3 = st.columns([2, 2, 3])
        with col_dl1:
            st.download_button(
                label=f"⬇️  Télécharger {ofx_name}",
                data=ofx_bytes,
                file_name=ofx_name,
                mime="application/x-ofx",
                key=f"dl_{file_key}",
                use_container_width=True,
            )
        with col_dl2:
            if excel_ok:
                st.download_button(
                    label=f"📊  Exporter Excel",
                    data=xlsx_bytes,
                    file_name=xlsx_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_xlsx_{file_key}",
                    use_container_width=True,
                    help="Tableau Excel professionnel avec résumé et mise en forme automatique",
                )
        with col_dl3:
            st.markdown(f"""
            <div style="padding:0.6rem 0; font-size:0.83rem; color:#7489b0; line-height:1.7;">
              <b style="color:#0f2d6b">Format :</b> OFX Standard ({target})
              &nbsp;&nbsp;·&nbsp;&nbsp;
              <b style="color:#0f2d6b">Devise :</b> {currency}
              &nbsp;&nbsp;·&nbsp;&nbsp;
              <b style="color:#0f2d6b">{len(txns_for_ofx)} transactions</b>
              {'&nbsp;&nbsp;·&nbsp;&nbsp;<span style="color:#d97706;font-weight:600">✏️ Données corrigées</span>' if has_changes else ''}
              &nbsp;&nbsp;·&nbsp;&nbsp;
              Encodage Latin-1 (Cegid / MyUnisoft)
            </div>
            """, unsafe_allow_html=True)

        # ── Panneau Diagnostic (toujours visible) ────────────────────────────
        with st.expander("🔧 Informations de diagnostic", expanded=mode_debug):
            ocr_status_lines = [
                f"**Mode OCR actif :** {'EasyOCR' if _EASYOCR_OK else 'Tesseract' if _TESSERACT_OK else 'Aucun (PDF texte uniquement)'}",
                f"**EasyOCR :** {'✅' if _EASYOCR_OK else '❌'}  |  **Tesseract :** {'✅' if _TESSERACT_OK else '❌'}  |  "
                f"**pdfplumber :** {'✅' if _PDFPLUMBER_OK else '❌'}  |  **pdf2image :** {'✅' if _PDF2IMAGE_OK else '❌'}",
            ]
            if ocr_mode:
                ocr_status_lines.append(f"**OCR utilisé pour ce fichier :** `{ocr_mode}`")
            st.markdown('\n\n'.join(ocr_status_lines))
            st.divider()

            col_dbg1, col_dbg2 = st.columns(2)
            with col_dbg1:
                st.markdown("**Données extraites du PDF**")
                dbg_info = {
                    'Banque détectée': bank,
                    'IBAN': info.get('iban') or '—',
                    'Code banque': info.get('_rib_bank') or '—',
                    'Guichet': info.get('_rib_agency') or '—',
                    'N° compte': info.get('_rib_account') or '—',
                    'Clé RIB': info.get('_rib_key') or '—',
                    'Période': f"{info.get('period_start','')} → {info.get('period_end','')}",
                    'Solde ouverture': fmt_amount(info.get('balance_open', 0.0), currency),
                    'Solde clôture': fmt_amount(info.get('balance_close', 0.0), currency),
                }
                for k, v in dbg_info.items():
                    st.markdown(f"- **{k} :** `{v}`")
            with col_dbg2:
                st.markdown("**Transactions extraites**")
                st.markdown(f"- **Total :** {len(txns_for_ofx)}")
                st.markdown(f"- **Débits :** {sum(1 for t in txns_for_ofx if t['type']=='DEBIT')}")
                st.markdown(f"- **Crédits :** {sum(1 for t in txns_for_ofx if t['type']=='CREDIT')}")
                if txns_for_ofx:
                    dates = [t['date'] for t in txns_for_ofx]
                    st.markdown(f"- **1ère date :** `{min(dates)}`")
                    st.markdown(f"- **Dernière date :** `{max(dates)}`")
                    st.markdown("**3 premières transactions brutes :**")
                    for t in txns_for_ofx[:3]:
                        st.code(f"{t['date']} | {t['type']:6s} | {t['amount']:>14.2f} | {t['name'][:40]}")

            # ── Texte extrait par page (mode debug uniquement) ───────────────
            if mode_debug:
                st.divider()
                st.markdown("**📄 Texte extrait du PDF par page** *(visible en mode debug)*")
                st.markdown(
                    "_Le texte ci-dessous est celui que voient les parseurs. "
                    "Utile pour diagnostiquer une extraction incorrecte._"
                )
                # Re-lire le texte directement depuis les bytes du fichier courant
                try:
                    import io as _io
                    import pdfplumber as _plumber
                    with _plumber.open(_io.BytesIO(file_bytes)) as _pdf:
                        for _pi, _page in enumerate(_pdf.pages, 1):
                            _txt = _page.extract_text() or ''
                            if _txt.strip():
                                st.text_area(
                                    f"Page {_pi}",
                                    _txt[:3000] + ('…' if len(_txt) > 3000 else ''),
                                    height=180,
                                    key=f"dbg_page_{file_key}_{_pi}"
                                )
                            else:
                                st.markdown(f"*Page {_pi} : aucun texte extractible (scan ?)*")
                except Exception as _e:
                    st.markdown(f"_Impossible de relire le texte : {_e}_")

        st.markdown('</div>', unsafe_allow_html=True)  # .result-card

    # ── Footer ────────────────────────────────────────────────────────────────
    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
    st.markdown("""
    <div class="footer-bar">
      <span>💱 <span class="highlight">OFX Bridge</span> v2.4</span>
      <span>🔒 Traitement 100 % local — Aucune donnée envoyée vers un serveur externe</span>
      <span>✉️ Compatible Quadra · MyUnisoft · Sage · EBP · Excel</span>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
